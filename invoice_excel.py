"""INVOICE 엑셀 출력 (STI AD USA INC 양식)"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quotation_excel import _logo

GRAY = PatternFill("solid", fgColor="D9DDE3")
LGRAY = PatternFill("solid", fgColor="EDEFF2")
DARK = "44546A"
THIN = Side(style="thin")
MED = Side(style="medium")
USD = '#,##0.00'

CO_LINES = ["STI AD USA INC",
            "2261 Gattis School Rd, Unit 100",
            "Round Rock, TX 78717",
            "512-999-4442",
            "Jaden@sti.co.kr"]

DEFAULT_BANK = ("Routing Number → 021000089\n"
                "Account Number → 31335874\n"
                "Account Name → STI AD USA\n"
                "Bank Name → CITIBANK N.A.\n"
                "City/State → NEW YORK, NY")


def _c(ws, r, c, v=None, *, bold=False, size=10, align="left", fill=None,
       color=None, fmt=None, italic=False, border=None):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    cell.font = Font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border
    return cell


def _band(ws, r, c1, c2, fill=LGRAY):
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).fill = fill


def build_invoice_excel(inv: dict, lines: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    for i, w in enumerate([28, 14, 13, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 로고 / INVOICE ----
    _logo(ws, "A1", height=40)
    _c(ws, 1, 5, "INVOICE", bold=False, size=26, align="right", color=DARK)
    ws.merge_cells("E1:F1")
    ws.row_dimensions[1].height = 34

    # ---- 회사 정보 / 번호 ----
    r = 3
    for i, line in enumerate(CO_LINES):
        _c(ws, r + i, 1, line, bold=(i == 0), size=10 if i == 0 else 9)
    _c(ws, 3, 4, "INVOICE NO.", bold=True, size=9)
    _c(ws, 3, 5, inv.get("invoice_no") or "", size=9)
    _c(ws, 4, 4, "DATE", bold=True, size=9)
    _c(ws, 4, 5, str(inv.get("invoice_date") or "")[:10], size=9)
    _c(ws, 5, 4, "PO No", bold=True, size=9)
    _c(ws, 5, 5, inv.get("po_no") or "", size=9)

    # ---- TO ----
    r = 9
    _c(ws, r, 1, "TO", bold=True, size=9)
    _c(ws, r + 1, 2, inv.get("client_name") or "", size=10)
    _c(ws, r + 2, 2, inv.get("client_address") or "", size=9)
    ws.merge_cells(start_row=r + 2, start_column=2, end_row=r + 2,
                   end_column=5)

    # ---- JOB / PAYMENT TERMS ----
    r = 14
    top = Border(top=MED)
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = top
    _c(ws, r, 1, "JOB", bold=True, size=9)
    _c(ws, r, 4, "PAYMENT TERMS", bold=True, size=9)
    _band(ws, r + 1, 1, 6, LGRAY)
    r += 2
    _band(ws, r, 1, 6, GRAY)
    _c(ws, r, 1, inv.get("job") or "", size=9, fill=GRAY)
    _c(ws, r, 4, inv.get("payment_terms") or "", size=9, fill=GRAY)

    # ---- 품목 ----
    r += 2
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = Border(bottom=THIN)
    _c(ws, r, 1, "DESCRIPTION", bold=True, size=9)
    _c(ws, r, 3, "QUANTITY", bold=True, size=9, align="right")
    _c(ws, r, 4, "AMOUNT", bold=True, size=9, align="right")
    _c(ws, r, 6, "TOTAL", bold=True, size=9, align="right")
    r += 1
    first_item = r
    cur_total = 0.0
    for i, l in enumerate(lines):
        fill = GRAY if i % 2 == 0 else None
        if fill:
            _band(ws, r, 1, 6, fill)
        qty = float(l.get("quantity") or 0)
        price = float(l.get("unit_price") or 0)
        _c(ws, r, 1, l.get("description") or "", size=9, fill=fill)
        _c(ws, r, 3, qty, size=9, align="right", fill=fill, fmt='#,##0.00')
        _c(ws, r, 4, price, size=9, align="right", fill=fill, fmt=USD)
        _c(ws, r, 6, f"=C{r}*D{r}", size=9, align="right", fill=fill,
           fmt=USD, bold=True)
        cur_total += qty * price
        r += 1

    # ---- 1) Position / 2) Period ----
    r += 1
    _c(ws, r, 1, "1) Position", size=9, bold=True)
    r += 1
    _band(ws, r, 1, 6, GRAY)
    _c(ws, r, 1, inv.get("position_note") or "", size=9, fill=GRAY)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    _c(ws, r, 1, "2) Period", size=9, bold=True, fill=LGRAY)
    _band(ws, r, 1, 6, LGRAY)
    r += 1
    _c(ws, r, 1, inv.get("period_note") or "", size=9)

    # ---- 3) 신청 내역 ----
    r += 2
    _c(ws, r, 1, "3) 신청 내역", size=9, bold=True, fill=LGRAY)
    _band(ws, r, 1, 6, LGRAY)
    r += 1
    box = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    heads = ["계약금액", "금회 신청금액", "이전 신청 금액", "잔여금액"]
    cols = [(1, 2), (3, 3), (4, 4), (5, 6)]
    for (c1, c2), h in zip(cols, heads):
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r,
                           end_column=c2)
        _c(ws, r, c1, h, bold=True, size=9, align="center", border=box)
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = box
    r += 1
    contract = float(inv.get("contract_amount") or 0)
    prev = float(inv.get("prev_billed") or 0)
    vals = [contract, cur_total, prev, contract - prev - cur_total]
    for (c1, c2), v in zip(cols, vals):
        if c2 > c1:
            ws.merge_cells(start_row=r, start_column=c1, end_row=r,
                           end_column=c2)
        _c(ws, r, c1, v, size=9, align="center", fill=GRAY, fmt=USD,
           bold=(c1 == 5), border=box)
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = box
            ws.cell(row=r, column=c).fill = GRAY

    # ---- TOTAL DUE ----
    r += 2
    for c in range(4, 7):
        ws.cell(row=r, column=c).border = Border(top=MED, bottom=MED)
    _c(ws, r, 4, "TOTAL DUE", bold=True, size=10, align="right")
    _c(ws, r, 6, contract - prev, bold=True, size=10, align="right", fmt=USD)

    # ---- 은행 정보 ----
    r += 2
    bank = inv.get("bank_info") or DEFAULT_BANK
    for line in bank.split("\n"):
        _c(ws, r, 1, line, size=8, italic=True)
        ws.cell(row=r, column=1).border = Border(top=THIN if line == bank.split("\n")[0] else None)
        r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
