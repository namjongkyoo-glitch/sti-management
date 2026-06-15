"""Admin 페이지: 사용자 관리 + 페이지별 접근 권한 설정"""
import os
import streamlit as st
import auth
from db import get_db


def render():
    if not auth.is_admin():
        st.error("관리자만 접근할 수 있습니다.")
        return

    helpers.page_title("Admin - 사용자 및 권한 관리")
    db = get_db()
    tab1, tab2, tab3, tab4 = st.tabs(["사용자 관리", "페이지 접근 권한",
                                      "📥 데이터 일괄 입력", "💾 백업 / 복원"])
    with tab3:
        bulk_import_tab(db)
    with tab4:
        backup_tab(db)

    # ------------------------------------------------------
    # 탭 1: 사용자 관리
    # ------------------------------------------------------
    with tab1:
        users = (db.table("app_users").select("*")
                 .order("created_at").execute().data)

        st.subheader("사용자 목록")
        for u in users:
            cols = st.columns([2, 2, 2, 1.5, 1.5, 1.5])
            cols[0].write(f"**{u['name']}**")
            cols[1].write(u["login_id"])
            cols[2].write(u.get("email") or "-")
            cols[3].write(auth.ROLE_LABEL.get(u["role"], u["role"]))
            cols[4].write("✅ 활성" if u["is_active"] else "⛔ 비활성")
            with cols[5].popover("관리"):
                new_pw = st.text_input("새 비밀번호", type="password",
                                       key=f"pw_{u['id']}")
                if st.button("비밀번호 변경", key=f"pwbtn_{u['id']}"):
                    if len(new_pw) < 6:
                        st.error("6자 이상")
                    else:
                        auth.reset_password(u["id"], new_pw)
                        st.success("변경 완료")
                if st.button("2FA 초기화 (기기 분실 시)", key=f"tf_{u['id']}"):
                    auth.reset_totp(u["id"])
                    st.success("2단계 인증이 초기화되었습니다. "
                               "다음 로그인 시 다시 등록합니다.")
                toggle_label = "비활성화" if u["is_active"] else "활성화"
                if st.button(toggle_label, key=f"tg_{u['id']}"):
                    db.table("app_users").update(
                        {"is_active": not u["is_active"]}
                    ).eq("id", u["id"]).execute()
                    st.rerun()

        st.divider()
        st.subheader("새 사용자 추가")
        with st.form("add_user", clear_on_submit=True):
            c1, c2 = st.columns(2)
            login_id = c1.text_input("로그인 ID *")
            name = c2.text_input("이름 *")
            email = c1.text_input("이메일")
            role = c2.selectbox(
                "권한", ["user", "viewer", "account", "admin"],
                format_func=lambda r: auth.ROLE_LABEL.get(r, r),
                help="Admin: 전체 / Account: Admin페이지 제외 전체(자금·직원·급여 포함, 편집 가능) / "
                     "관찰자: Admin페이지 제외 전체 조회만 가능(편집 불가) / "
                     "일반: 매트릭스 설정 페이지만 (자금·직원·급여·Admin 불가)")
            pw = c1.text_input("초기 비밀번호 *", type="password")
            ok = st.form_submit_button("사용자 추가", type="primary")
        if ok:
            if not login_id or not name or len(pw) < 6:
                st.error("ID, 이름은 필수이며 비밀번호는 6자 이상입니다.")
            else:
                try:
                    auth.create_user(login_id, name, pw, role=role, email=email)
                    st.success(f"'{name}' 사용자가 추가되었습니다.")
                    st.rerun()
                except Exception as e:
                    st.error(f"추가 실패 (ID 중복 여부 확인): {e}")

    # ------------------------------------------------------
    # 탭 2: 페이지 접근 권한
    # ------------------------------------------------------
    with tab2:
        st.caption("Admin은 전체, Account는 Admin 페이지를 제외한 전체에 자동 접근됩니다. "
                   "아래 매트릭스는 '일반' 사용자에게만 적용되며, "
                   "자금 집행/자금 현황/직원/급여/Admin 페이지는 일반 사용자에게 항상 차단됩니다.")
        normal_users = [u for u in users if u["role"] == "user"]
        if not normal_users:
            st.info("일반 사용자가 없습니다.")
            return

        target = st.selectbox("사용자 선택", normal_users,
                              format_func=lambda u: f"{u['name']} ({u['login_id']})")
        blocked = auth.RESTRICTED_PAGES | auth.ADMIN_ONLY_PAGES
        pages = [p for p in auth.get_all_pages() if p["code"] not in blocked]
        existing = (db.table("page_permissions").select("*")
                    .eq("user_id", target["id"]).execute().data)
        emap = {r["page_code"]: r for r in existing}

        st.write("")
        h = st.columns([3, 1.5, 1.5])
        h[0].markdown("**페이지**")
        h[1].markdown("**보기**")
        h[2].markdown("**편집**")

        new_perms = {}
        for p in pages:
            row = st.columns([3, 1.5, 1.5])
            row[0].write(p["name_kr"])
            cur = emap.get(p["code"], {})
            v = row[1].checkbox(" ", value=cur.get("can_view", False),
                                key=f"v_{target['id']}_{p['code']}",
                                label_visibility="collapsed")
            e = row[2].checkbox("  ", value=cur.get("can_edit", False),
                                key=f"e_{target['id']}_{p['code']}",
                                label_visibility="collapsed")
            new_perms[p["code"]] = (v, e)

        if st.button("권한 저장", type="primary"):
            # 기존 권한 삭제 후 재삽입
            db.table("page_permissions").delete().eq(
                "user_id", target["id"]).execute()
            rows = [{"user_id": target["id"], "page_code": code,
                     "can_view": v, "can_edit": e}
                    for code, (v, e) in new_perms.items() if v or e]
            if rows:
                db.table("page_permissions").insert(rows).execute()
            st.success("권한이 저장되었습니다.")


# ============================================================
# 탭 3: 데이터 일괄 입력 (엑셀 템플릿 다운로드 -> 업로드)
# ============================================================
import pandas as pd
from io import BytesIO
import helpers

T_SHEETS = {
    "고객사": ["고객사명*", "담당자", "전화", "이메일", "주소", "비고"],
    "협력업체": ["업체명*", "공종", "담당자", "전화", "이메일", "주소",
              "은행명", "계좌번호", "Routing", "Zelle", "비고"],
    "직원": ["이름*", "영문이름", "직책", "입사일(YYYY-MM-DD)", "전화", "이메일",
           "기본급", "일비", "OT단가", "중식지원", "차량지원", "통신지원",
           "송금수수료", "은행명", "계좌번호", "Routing", "계좌종류", "Zelle"],
    "프로젝트": ["코드(빈칸=자동)", "프로젝트명*", "고객사", "현장위치",
             "계약금액", "시작일(YYYY-MM-DD)", "종료일", "PM",
             "상태(진행중/완료/보류)"],
    "통장": ["은행명*", "별칭(용도)", "계좌번호", "기초금액"],
    "대출": ["대출기관*", "대출명", "원금*", "연이율(%)",
           "실행일(YYYY-MM-DD)", "만기일", "매월상환일"],
    "거래내역": ["일자(YYYY-MM-DD)*", "구분(수입/지출)*", "프로젝트코드",
             "계정명", "거래처(업체명)", "통장(은행명)", "금액*", "내용"],
    "견적품의": ["프로젝트코드*", "프로젝트명", "고객사",
             "버전(본품의/1차변경/2차변경..)*", "변경사유",
             "수주금액", "재료비", "외주비", "직접경비",
             "노무비", "제조간접경비", "판관비", "예비비",
             "승인일(YYYY-MM-DD)", "최종버전(Y/N)"],
}

T_SAMPLES = {
    "고객사": ["SECAI", "홍길동", "512-000-0000", "a@b.com", "Taylor, TX", ""],
    "협력업체": ["HANWOOL", "배관 (Piping)", "김담당", "", "", "",
              "Chase", "123456789", "021000021", "", ""],
    "직원": ["남종규", "Jaden Nam", "Branch Manager", "2022-09-01", "", "",
           10465, 0, 0, 400, 0, 0, 15, "Chase", "", "", "Checking", ""],
    "프로젝트": ["", "Taylor Fab1 Cycle Purge", "SECAI", "Taylor, TX",
             130000, "2026-04-01", "2026-04-30", "우성진", "진행중"],
    "통장": ["Chase", "운영", "123456789", 50000],
    "대출": ["본사(STI HQ)", "운영자금 대여", 2900000, 5.0,
           "2023-01-01", "2027-12-31", 25],
    "거래내역": ["2026-04-30", "수입", "P2026-001", "기성금/계약수입",
             "", "Chase", 50000, "1차 기성"],
    "견적품의": ["P2026-001", "Taylor Fab1 Cycle Purge", "SECAI",
             "본품의", "", 130000, 0, 100000, 0, 3900, 0, 5200, 0,
             "2026-04-17", "N"],
}


def _g(row, col, default=""):
    v = row.get(col)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return default
    return v


def _s(row, col):
    v = _g(row, col, "")
    if v == "":
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    return "" if s.lower() in ("nat", "nan", "none") else s


def _n(row, col):
    try:
        v = _g(row, col, 0)
        return float(v) if str(v).strip() != "" else 0.0
    except Exception:
        return 0.0


def _d(row, col):
    s = _s(row, col)
    if not s or s.lower() in ("nat", "nan", "none", "null"):
        return None
    # 'YYYY-MM-DD HH:MM:SS' 또는 'YYYY-MM-DD' 형태에서 날짜만
    s = s.replace("/", "-")[:10]
    # 유효성 간이 확인 (YYYY-MM-DD)
    parts = s.split("-")
    if len(parts) == 3 and parts[0].isdigit() and len(parts[0]) == 4:
        return s
    return None


def build_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    wb.remove(wb.active)
    for name, cols in T_SHEETS.items():
        ws = wb.create_sheet(name)
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
            ws.column_dimensions[cell.column_letter].width = \
                max(12, len(c) + 4)
        for i, v in enumerate(T_SAMPLES[name], 1):
            ws.cell(row=2, column=i, value=v)
        ws.cell(row=4, column=1,
                value="※ 2행의 예시는 지우고 실제 데이터를 입력하세요. "
                      "*표시는 필수 항목입니다.")
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def account_id_any(name):
    """계정명 -> id (세부계정 우선, 없으면 중분류)"""
    name = name.strip()
    for lv in (3, 2):
        for a in helpers.load_accounts():
            if a["level"] == lv and a["name_kr"] == name:
                return a["id"]
    return None


def bulk_import_tab(db):
    st.markdown("**① 템플릿 다운로드 → ② 기존 자료 입력 → ③ 업로드 → ④ 가져오기**")
    st.download_button("⬇️ 일괄 입력 템플릿 다운로드 (엑셀)",
                       data=build_template(),
                       file_name="STI_일괄입력_템플릿.xlsx",
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")
    st.caption("시트 구성: 고객사 / 협력업체 / 직원 / 프로젝트 / 통장 / 대출 / 거래내역. "
               "필요한 시트만 채워도 됩니다. 이름이 같은 항목은 중복 등록을 건너뜁니다. "
               "거래내역은 재업로드 시 중복될 수 있으니 한 번만 올리세요.")

    # ---- 현재 등록 데이터 참조표 (입력 시 코드/이름 맞추기용) ----
    st.download_button("📑 현재 등록 데이터 참조표 다운로드 (코드/ID 목록)",
                       data=build_reference(db),
                       file_name="STI_참조표_코드ID.xlsx",
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")
    st.caption("거래내역/견적품의 입력 시 사용할 프로젝트코드, 계정명, "
               "협력업체명, 통장명, 고객사명을 확인하세요.")

    up = st.file_uploader("작성한 엑셀 업로드", type=["xlsx"])
    if not up:
        return
    try:
        sheets = pd.read_excel(up, sheet_name=None)
    except Exception as e:
        st.error(f"엑셀을 읽을 수 없습니다: {e}")
        return

    for name in T_SHEETS:
        if name not in sheets:
            continue
        df = sheets[name].dropna(how="all")
        # 안내문 행 제거
        df = df[df[df.columns[0]].astype(str).str.startswith("※") == False]
        if df.empty:
            continue
        with st.expander(f"📄 {name} — {len(df)}행", expanded=False):
            st.dataframe(df.head(20), use_container_width=True,
                         hide_index=True)
            if st.button(f"'{name}' 가져오기", key=f"imp_{name}",
                         type="primary"):
                ok, skip, errs = import_sheet(db, name, df)
                helpers.clear_caches()
                st.success(f"{name}: {ok}건 등록, {skip}건 건너뜀(중복 등)")
                for e in errs[:10]:
                    st.warning(e)


def import_sheet(db, name, df):
    ok = skip = 0
    errs = []
    if name == "고객사":
        existing = {c["name"] for c in
                    db.table("clients").select("name").execute().data}
        for _, r in df.iterrows():
            nm = _s(r, "고객사명*")
            if not nm or nm in existing:
                skip += 1
                continue
            db.table("clients").insert({
                "name": nm, "contact": _s(r, "담당자"), "phone": _s(r, "전화"),
                "email": _s(r, "이메일"), "address": _s(r, "주소"),
                "notes": _s(r, "비고")}).execute()
            existing.add(nm)
            ok += 1

    elif name == "협력업체":
        existing = {v["name"] for v in
                    db.table("vendors").select("name").execute().data}
        for _, r in df.iterrows():
            nm = _s(r, "업체명*")
            if not nm or nm in existing:
                skip += 1
                continue
            trade = _s(r, "공종")
            if trade:
                helpers.add_trade(trade)
            db.table("vendors").insert({
                "name": nm, "trade": trade, "contact": _s(r, "담당자"),
                "phone": _s(r, "전화"), "email": _s(r, "이메일"),
                "address": _s(r, "주소"), "bank_name": _s(r, "은행명"),
                "account_no": _s(r, "계좌번호"), "routing_no": _s(r, "Routing"),
                "zelle": _s(r, "Zelle"), "notes": _s(r, "비고")}).execute()
            existing.add(nm)
            ok += 1

    elif name == "직원":
        existing = {e["name"] for e in
                    db.table("employees").select("name").execute().data}
        for _, r in df.iterrows():
            nm = _s(r, "이름*")
            if not nm or nm in existing:
                skip += 1
                continue
            db.table("employees").insert({
                "name": nm, "name_en": _s(r, "영문이름"),
                "position": _s(r, "직책"),
                "hire_date": _d(r, "입사일(YYYY-MM-DD)"),
                "phone": _s(r, "전화"), "email": _s(r, "이메일"),
                "base_salary": _n(r, "기본급"), "per_diem": _n(r, "일비"),
                "ot_rate": _n(r, "OT단가"), "meal_support": _n(r, "중식지원"),
                "vehicle_support": _n(r, "차량지원"),
                "telecom_support": _n(r, "통신지원"),
                "transfer_fee": _n(r, "송금수수료") or 15,
                "bank_name": _s(r, "은행명"), "account_no": _s(r, "계좌번호"),
                "routing_no": _s(r, "Routing"),
                "account_type": _s(r, "계좌종류"),
                "zelle": _s(r, "Zelle")}).execute()
            existing.add(nm)
            ok += 1

    elif name == "프로젝트":
        existing = {p["code"] for p in
                    db.table("projects").select("code").execute().data}
        names = {p["name"] for p in
                 db.table("projects").select("name").execute().data}
        for _, r in df.iterrows():
            pname = _s(r, "프로젝트명*")
            if not pname or pname in names:
                skip += 1
                continue
            code = _s(r, "코드(빈칸=자동)") or \
                helpers.next_no("projects", "code", "P")
            if code in existing:
                skip += 1
                errs.append(f"프로젝트 코드 중복: {code}")
                continue
            db.table("projects").insert({
                "code": code, "name": pname, "client": _s(r, "고객사"),
                "location": _s(r, "현장위치"),
                "contract_amount": _n(r, "계약금액"),
                "start_date": _d(r, "시작일(YYYY-MM-DD)"),
                "end_date": _d(r, "종료일"), "pm": _s(r, "PM"),
                "status": _s(r, "상태(진행중/완료/보류)") or "진행중",
            }).execute()
            existing.add(code)
            names.add(pname)
            ok += 1

    elif name == "통장":
        existing = {(b["bank_name"], b.get("account_no") or "") for b in
                    db.table("bank_accounts")
                    .select("bank_name,account_no").execute().data}
        for _, r in df.iterrows():
            bn = _s(r, "은행명*")
            no = _s(r, "계좌번호")
            if not bn or (bn, no) in existing:
                skip += 1
                continue
            db.table("bank_accounts").insert({
                "bank_name": bn, "account_name": _s(r, "별칭(용도)"),
                "account_no": no,
                "opening_balance": _n(r, "기초금액")}).execute()
            existing.add((bn, no))
            ok += 1

    elif name == "대출":
        for _, r in df.iterrows():
            lender = _s(r, "대출기관*")
            principal = _n(r, "원금*")
            if not lender or principal <= 0:
                skip += 1
                continue
            db.table("loans").insert({
                "lender": lender, "loan_name": _s(r, "대출명"),
                "principal": principal,
                "interest_rate": _n(r, "연이율(%)"),
                "start_date": _d(r, "실행일(YYYY-MM-DD)"),
                "maturity_date": _d(r, "만기일"),
                "payment_day": int(_n(r, "매월상환일")) or None}).execute()
            ok += 1

    elif name == "견적품의":
        return import_estimates(db, df)

    elif name == "거래내역":
        prj_map = {p["code"]: p["id"] for p in
                   db.table("projects").select("id,code").execute().data}
        bank_map = {}
        for b in db.table("bank_accounts").select("*").execute().data:
            bank_map[b["bank_name"]] = b["id"]
            if b.get("account_name"):
                bank_map[f"{b['bank_name']} {b['account_name']}"] = b["id"]
        rows = []
        for idx, r in df.iterrows():
            tx_date = _d(r, "일자(YYYY-MM-DD)*")
            tx_type = _s(r, "구분(수입/지출)*")
            amt = _n(r, "금액*")
            if not tx_date or tx_type not in ("수입", "지출") or amt <= 0:
                skip += 1
                errs.append(f"거래내역 {idx+2}행: 일자/구분/금액 확인 필요")
                continue
            pcode = _s(r, "프로젝트코드")
            pid = prj_map.get(pcode)
            if pcode and pid is None:
                errs.append(f"거래내역 {idx+2}행: 프로젝트코드 '{pcode}' 없음 "
                            "(프로젝트 없이 등록됨)")
            acc_name = _s(r, "계정명")
            acc_id = account_id_any(acc_name) if acc_name else None
            if acc_name and acc_id is None:
                errs.append(f"거래내역 {idx+2}행: 계정 '{acc_name}' 미일치 "
                            "(계정 없이 등록됨)")
            rows.append({
                "tx_type": tx_type, "tx_date": tx_date, "amount": amt,
                "project_id": pid, "account_id": acc_id,
                "vendor_id": helpers.vendor_id_by_name(_s(r, "거래처(업체명)")),
                "bank_account_id": bank_map.get(_s(r, "통장(은행명)")),
                "description": _s(r, "내용"),
            })
        if rows:
            db.table("transactions").insert(rows).execute()
            ok = len(rows)
    return ok, skip, errs


# ============================================================
# 과거 견적/품의 일괄 등록 (버전별 행 -> 견적/버전/품의/예산 생성)
# ============================================================
DIRECT_MAP = {"재료비": "원재료", "외주비": "외주비", "직접경비": "직접경비"}
INDIRECT_MAP = {"노무비": "노무비", "제조간접경비": "제조간접경비", "판관비": "판관비"}


def import_estimates(db, df):
    ok = skip = 0
    errs = []
    prj_map = {p["code"]: p for p in
               db.table("projects").select("*").execute().data}
    groups, order = {}, []
    for _, r in df.iterrows():
        code = _s(r, "프로젝트코드*")
        if not code:
            continue
        if code not in groups:
            groups[code] = []
            order.append(code)
        groups[code].append(r)

    for code in order:
        rowset = groups[code]
        prj = prj_map.get(code)
        first = rowset[0]
        pname = _s(first, "프로젝트명") or code
        client = _s(first, "고객사")

        ests = db.table("estimates").select("*").eq("title", pname).execute().data
        if ests:
            est = ests[0]
        else:
            est = db.table("estimates").insert({
                "estimate_no": helpers.next_no("estimates", "estimate_no", "Q"),
                "title": pname, "client": client, "status": "수주",
            }).execute().data[0]

        last_row, vno = None, 0
        for r in rowset:
            vno += 1
            vlabel = _s(r, "버전(본품의/1차변경/2차변경..)*") or ("v%d" % vno)
            order_amt = _n(r, "수주금액")
            db.table("estimate_versions").insert({
                "estimate_id": est["id"], "version_no": vno,
                "version_label": vlabel, "order_amount": order_amt,
                "change_reason": _s(r, "변경사유"),
            }).execute()
            db.table("proposals").insert({
                "doc_no": "과거-%s-%d" % (code, vno),
                "proposal_type": "본품의" if vno == 1 else "변경품의",
                "title": "%s - %s" % (pname, vlabel),
                "project_name": pname, "client": client,
                "order_amount": order_amt,
                "material_cost": _n(r, "재료비"),
                "outsourcing_cost": _n(r, "외주비"),
                "direct_expense": _n(r, "직접경비"),
                "labor_cost": _n(r, "노무비"),
                "mfg_overhead": _n(r, "제조간접경비"),
                "sga_cost": _n(r, "판관비"),
                "reserve": _n(r, "예비비"),
                "status": "승인",
                "result_note": _s(r, "변경사유"),
                "decided_at": _d(r, "승인일(YYYY-MM-DD)"),
            }).execute()
            last_row = r
            ok += 1

        final_row = None
        for r in rowset:
            if _s(r, "최종버전(Y/N)").upper() == "Y":
                final_row = r
        if final_row is None:
            final_row = last_row

        if prj is None:
            prj = db.table("projects").insert({
                "code": code, "name": pname, "client": client,
                "contract_amount": _n(final_row, "수주금액"),
                "status": "진행중",
            }).execute().data[0]
            prj_map[code] = prj
        else:
            db.table("projects").update({
                "contract_amount": _n(final_row, "수주금액"),
            }).eq("id", prj["id"]).execute()

        db.table("budget_lines").delete().eq("project_id", prj["id"]).execute()
        brows = []
        allmap = {}
        allmap.update(DIRECT_MAP)
        allmap.update(INDIRECT_MAP)
        for col, mid in allmap.items():
            amt = _n(final_row, col)
            if amt > 0:
                aid = account_id_any(mid)
                if aid:
                    brows.append({
                        "project_id": prj["id"], "account_id": aid,
                        "amount": amt, "notes": "과거자료 일괄등록",
                    })
        if brows:
            db.table("budget_lines").insert(brows).execute()

    return ok, skip, errs


# ============================================================
# 현재 등록 데이터 참조표 (코드/ID/이름) — 일괄 입력 보조
# ============================================================
def build_reference(db):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="DDEBF7")

    def add_sheet(name, cols, rows):
        ws = wb.create_sheet(name)
        for i, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.font = Font(bold=True)
            cell.fill = head_fill
            ws.column_dimensions[cell.column_letter].width = max(12, len(str(c)) + 4)
        for r, row in enumerate(rows, 2):
            for i, v in enumerate(row, 1):
                ws.cell(row=r, column=i, value=v)

    # 프로젝트
    prjs = (db.table("projects").select("*")
            .order("code").execute().data)
    add_sheet("프로젝트", ["ID", "프로젝트코드", "프로젝트명", "고객사",
                       "계약금액", "상태"],
              [[p["id"], p["code"], p["name"], p.get("client") or "",
                float(p.get("contract_amount") or 0), p.get("status") or ""]
               for p in prjs if not p.get("is_common_pool")])

    # 계정 (거래내역 계정명 입력용)
    accs = helpers.load_accounts()
    add_sheet("계정", ["ID", "계정코드", "계정명", "레벨", "구분"],
              [[a["id"], a.get("code") or "", a["name_kr"], a["level"],
                "공통비" if a.get("is_common") else ""]
               for a in sorted(accs, key=lambda x: str(x.get("code") or ""))])

    # 협력업체
    vendors = db.table("vendors").select("*").order("name").execute().data
    add_sheet("협력업체", ["ID", "업체명", "공종"],
              [[v["id"], v["name"], v.get("trade") or ""] for v in vendors])

    # 고객사
    clients = db.table("clients").select("*").order("name").execute().data
    add_sheet("고객사", ["ID", "고객사명"],
              [[c["id"], c["name"]] for c in clients])

    # 통장
    banks = db.table("bank_accounts").select("*").order("bank_name").execute().data
    add_sheet("통장", ["ID", "은행명", "별칭", "계좌번호", "입력시_통장명"],
              [[b["id"], b["bank_name"], b.get("account_name") or "",
                b.get("account_no") or "", b["bank_name"]] for b in banks])

    # 직원
    emps = db.table("employees").select("*").order("name").execute().data
    add_sheet("직원", ["ID", "이름", "직책"],
              [[e["id"], e["name"], e.get("position") or ""] for e in emps])

    # 대출
    loans = db.table("loans").select("*").order("start_date").execute().data
    add_sheet("대출", ["ID", "기관", "대출명", "실행일", "원금"],
              [[l["id"], l["lender"], l.get("loan_name") or "",
                l.get("start_date") or "", float(l.get("principal") or 0)]
               for l in loans])

    # 견적/품의
    ests = db.table("estimates").select("*").order("estimate_no").execute().data
    add_sheet("견적", ["ID", "견적번호", "견적명", "고객사", "상태"],
              [[e["id"], e["estimate_no"], e["title"], e.get("client") or "",
                e.get("status") or ""] for e in ests])
    props = (db.table("proposals").select("*")
             .order("doc_no").execute().data)
    add_sheet("품의서", ["ID", "문서번호", "제목", "종류", "상태"],
              [[p["id"], p["doc_no"], p.get("title") or "",
                p.get("proposal_type") or "", p.get("status") or ""]
               for p in props])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 탭 4: 백업 / 복원 (Admin)
# ============================================================
def backup_tab(db):
    import backup_util
    from datetime import date

    st.markdown("### 💾 데이터 백업")
    st.caption("정기적으로 백업을 받아 안전한 곳에 보관하세요. "
               "엑셀은 보관·열람용, JSON은 복원용입니다.")

    c = st.columns(2)
    with c[0]:
        if st.button("📊 엑셀 백업 생성", use_container_width=True):
            with st.spinner("전체 데이터를 엑셀로 정리 중..."):
                st.session_state["_bk_xlsx"] = backup_util.backup_excel(db)
        if "_bk_xlsx" in st.session_state:
            st.download_button(
                "⬇️ 엑셀 백업 다운로드", data=st.session_state["_bk_xlsx"],
                file_name=f"STI_백업_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet", use_container_width=True)
    with c[1]:
        if st.button("🗄️ JSON 백업 생성 (복원용)", use_container_width=True):
            with st.spinner("전체 데이터를 JSON으로 추출 중..."):
                st.session_state["_bk_json"] = backup_util.backup_json(db)
        if "_bk_json" in st.session_state:
            st.download_button(
                "⬇️ JSON 백업 다운로드", data=st.session_state["_bk_json"],
                file_name=f"STI_백업_{date.today()}.json",
                mime="application/json", use_container_width=True)

    st.divider()
    st.markdown("### 🧩 프로그램(소스코드) 백업")
    st.caption("현재 실행 중인 프로그램 폴더의 소스 전체를 zip으로 받습니다.")
    app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if st.button("📦 소스코드 zip 생성", use_container_width=True):
        with st.spinner("소스 파일 압축 중..."):
            st.session_state["_bk_src"] = backup_util.backup_source(app_dir)
    if "_bk_src" in st.session_state:
        st.download_button(
            "⬇️ 소스코드 zip 다운로드", data=st.session_state["_bk_src"],
            file_name=f"STI_소스백업_{date.today()}.zip",
            mime="application/zip", use_container_width=True)

    st.divider()
    st.markdown("### ♻️ 복원 (JSON 업로드)")
    st.warning("복원은 신중히! 작업 전에 반드시 현재 상태를 먼저 백업하세요.")
    up = st.file_uploader("JSON 백업 파일 업로드", type=["json"])
    if up:
        mode = st.radio(
            "복원 방식",
            ["merge (기존 유지 + 백업으로 덮어쓰기/추가)",
             "replace (해당 테이블 전체 비우고 백업으로 교체)"],
            help="merge: id 기준 upsert (안전). "
                 "replace: 기존 데이터 삭제 후 교체 (위험).")
        mode_key = "merge" if mode.startswith("merge") else "replace"
        confirm = st.text_input(
            f"복원을 진행하려면 '{mode_key}' 를 입력하세요")
        if st.button("♻️ 복원 실행", type="primary"):
            if confirm.strip() != mode_key:
                st.error("확인 문구가 일치하지 않습니다.")
            else:
                with st.spinner("복원 중... (잠시 걸릴 수 있습니다)"):
                    try:
                        result = backup_util.restore_json(
                            db, up.getvalue(), mode=mode_key)
                        helpers.clear_caches()
                        st.success("복원이 완료되었습니다.")
                        st.json(result)
                    except Exception as e:
                        st.error(f"복원 중 오류: {e}")
