"""자금수지 통합 엑셀: Sheet1 요약(자금요약+PJT손익), Sheet2 월별 자금수지(연동 수식)"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BLUE = PatternFill("solid", fgColor="9DC3E6")
LBLUE = PatternFill("solid", fgColor="DDEBF7")
GREENF = PatternFill("solid", fgColor="E2EFDA")
GOLD = PatternFill("solid", fgColor="FFC000")
GRAYF = PatternFill("solid", fgColor="D9D9D9")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = '#,##0;[Red](#,##0)'


def _w(ws, r, c, v=None, *, fill=None, bold=False, size=9, align="center",
       fmt=None, color=None):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size, color=color)
    if fmt:
        cell.number_format = fmt
    return cell


def _fillrow(ws, r, c1, c2, fill=None):
    for c in range(c1, c2 + 1):
        ws.cell(row=r, column=c).border = BORDER
        if fill:
            ws.cell(row=r, column=c).fill = fill


# ============================================================
def sheet_summary(wb, s):
    ws = wb.active
    ws.title = "요약"
    for i, w in enumerate([6, 16, 14, 14, 14, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="미국법인 자금 요약 검토")
    t.font = Font(bold=True, size=16)
    t.alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=6, value="(단위: USD)").alignment = \
        Alignment(horizontal="right")

    r = 3
    hdr = ["구 분", "", "실적 누계", "계획 반영 후", "합 계", "비 고"]
    for i, h in enumerate(hdr, 1):
        _w(ws, r, i, h, fill=BLUE, bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    def row3(g, item, a, note="", fill=None, bold=False):
        nonlocal r
        _w(ws, r, 1, g, fill=fill, bold=bold)
        _w(ws, r, 2, item, fill=fill, bold=bold, align="left")
        _w(ws, r, 3, a, fill=fill, bold=bold, fmt=NUM, align="right")
        _w(ws, r, 4, "", fill=fill)
        _w(ws, r, 5, a, fill=fill, bold=bold, fmt=NUM, align="right")
        _w(ws, r, 6, note, fill=fill, align="left", size=8)
        r += 1

    in_sub = s["pjt_in"] + s["etc_in"] + s["loan_issue"]
    out_sub = s["direct_out"] + s["common_out"] + s["loan_repay"]
    row3("입금", "PJT 채권(기성 등)", s["pjt_in"])
    row3("", "자본금/기타 입금", s["etc_in"])
    row3("", "대출 실행", s["loan_issue"])
    row3("", "소 계", in_sub, fill=GREENF, bold=True)
    row3("지출", "PJT 직접비", s["direct_out"])
    row3("", "공통비", s["common_out"])
    row3("", "대출 상환(원금)", s["loan_repay"])
    row3("", "소 계", out_sub, fill=GREENF, bold=True)
    row3("잔액", "예금 잔액(통장 합계)", s["bank"], "통장 관리 기준")
    row3("", "대출 잔액", s["loan_balance"])
    row3("수지", "총 자금수지", in_sub - out_sub, fill=LBLUE, bold=True)

    # ---- PJT별 수주/직접비/손익 ----
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    t2 = ws.cell(row=r, column=1, value="PJT 별 수주 / 직접비 / 손익")
    t2.font = Font(bold=True, size=13)
    t2.alignment = Alignment(horizontal="center")
    r += 1
    cols = ["고객사", "프로젝트", "수주금액", "입금액", "미수금",
            "발주(직접비)", "기지급", "미지급", "공통비", "(예상)손익"]
    for i, h in enumerate(cols, 1):
        _w(ws, r, i, h, fill=BLUE, bold=True, size=8)
        ws.column_dimensions[get_column_letter(i)].width = \
            max(ws.column_dimensions[get_column_letter(i)].width or 10, 11)
    r += 1
    tot = {k: 0.0 for k in ["contract", "received", "receivable",
                            "direct_budget", "direct_paid", "direct_unpaid",
                            "common", "profit"]}
    for p in s["pjt_rows"]:
        _w(ws, r, 1, p["client"], size=8)
        _w(ws, r, 2, p["name"], size=8, align="left")
        for i, k in enumerate(["contract", "received", "receivable",
                               "direct_budget", "direct_paid",
                               "direct_unpaid", "common", "profit"], 3):
            _w(ws, r, i, p[k], fmt=NUM, align="right", size=8,
               color="FF0000" if (k == "profit" and p[k] < 0) else None)
            tot[k] += p[k]
        r += 1
    _w(ws, r, 1, "합계", fill=GREENF, bold=True)
    _w(ws, r, 2, "", fill=GREENF)
    for i, k in enumerate(["contract", "received", "receivable",
                           "direct_budget", "direct_paid", "direct_unpaid",
                           "common", "profit"], 3):
        _w(ws, r, i, tot[k], fill=GREENF, bold=True, fmt=NUM, align="right",
           size=8, color="FF0000" if (k == "profit" and tot[k] < 0) else None)


# ============================================================
def sheet_cashflow(wb, months, current_ym, opening, data, loan_bal_now):
    ws = wb.create_sheet("자금수지")
    n = len(months)
    # 컬럼: A 구분(8) B 항목(20) C 업체명(14) D.. 월(10) 합계(11)
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    for j in range(n):
        ws.column_dimensions[get_column_letter(4 + j)].width = 10
    tcol = 4 + n
    ws.column_dimensions[get_column_letter(tcol)].width = 12

    def L(j):
        return get_column_letter(4 + j)

    # 헤더
    _w(ws, 1, 1, f"기준일 : {current_ym}", fill=LBLUE, bold=True, align="left")
    _fillrow(ws, 1, 1, 3, LBLUE)
    r = 2
    _w(ws, r, 1, "구 분", fill=BLUE, bold=True)
    _w(ws, r, 2, "", fill=BLUE)
    _w(ws, r, 3, "업체명/항목", fill=BLUE, bold=True)
    for j, ym in enumerate(months):
        lab = f"{ym[2:4]}.{int(ym[5:7])}월" + ("" if ym <= current_ym else "(E)")
        _w(ws, r, 4 + j, lab, fill=BLUE, bold=True, size=8)
    _w(ws, r, tcol, "기간 합계", fill=BLUE, bold=True, size=8)
    r += 1

    def sumrow(row):
        _w(ws, row, tcol, f"=SUM({L(0)}{row}:{L(n-1)}{row})",
           fmt=NUM, align="right", bold=True)

    # ---- 기초 시재 (다음 달 기초 = 전월 기말, 수식 연동) ----
    r_base = r
    _w(ws, r, 1, "기초 시재", bold=True, fill=GRAYF)
    _fillrow(ws, r, 1, 3, GRAYF)
    _w(ws, r, 4, opening, fmt=NUM, align="right", bold=True, fill=GRAYF)
    # 기말 행 번호는 나중에 채움 -> placeholder, 마지막에 수식 입력
    r += 1

    def section(title, rows, three_col=False):
        """rows: [{'prj','label','vals'}] / 반환 (소계행)"""
        nonlocal r
        start = r
        for item in rows:
            _w(ws, r, 1, "")
            _w(ws, r, 2, item.get("prj", item.get("label", "")), size=8,
               align="left")
            _w(ws, r, 3, item.get("label", "") if three_col else "",
               size=8, align="left")
            for j in range(n):
                v = item["vals"][j]
                _w(ws, r, 4 + j, v if v else None, fmt=NUM, align="right",
                   size=8)
            sumrow(r)
            r += 1
        # 소계
        _w(ws, r, 2, "소 계", fill=GRAYF, bold=True)
        _fillrow(ws, r, 1, 3, GRAYF)
        for j in range(n):
            _w(ws, r, 4 + j, f"=SUM({L(j)}{start}:{L(j)}{r-1})",
               fill=GRAYF, bold=True, fmt=NUM, align="right", size=8)
        _w(ws, r, tcol, f"=SUM({L(0)}{r}:{L(n-1)}{r})",
           fill=GRAYF, bold=True, fmt=NUM, align="right")
        sub = r
        r += 1
        # 구분 라벨
        ws.merge_cells(start_row=start, start_column=1, end_row=sub - 1
                       if sub > start else start, end_column=1)
        _w(ws, start, 1, title, bold=True)
        return sub

    inc_rows = [{"prj": x["label"], "label": "", "vals": x["vals"]}
                for x in data["income"]] or \
        [{"prj": "(입금 계획 없음)", "label": "", "vals": [0] * n}]
    r_in = section("영업\n입금", inc_rows)

    exp_rows = [{"prj": x["prj"], "label": x["label"], "vals": x["vals"]}
                for x in data["expense"]] or \
        [{"prj": "(지출 계획 없음)", "label": "", "vals": [0] * n}]
    r_exp = section("영업지출\n매입대", exp_rows, three_col=True)

    com_rows = [{"prj": x["label"], "label": "", "vals": x["vals"]}
                for x in data["common"]] or \
        [{"prj": "(공통비 없음)", "label": "", "vals": [0] * n}]
    r_com = section("공통비", com_rows)

    # 영업활동 수지
    _w(ws, r, 1, "영업활동 자금수지", fill=GOLD, bold=True)
    _fillrow(ws, r, 1, 3, GOLD)
    for j in range(n):
        _w(ws, r, 4 + j, f"={L(j)}{r_in}-{L(j)}{r_exp}-{L(j)}{r_com}",
           fill=GOLD, bold=True, fmt=NUM, align="right", size=8)
    sumrow(r)
    ws.cell(row=r, column=tcol).fill = GOLD
    r_op = r
    r += 1

    # ---- 재무활동 ----
    fin_start = r
    fin_idx = {}
    for it, vals in data["fin"].items():
        _w(ws, r, 2, it, size=8, align="left")
        _w(ws, r, 3, "")
        for j in range(n):
            _w(ws, r, 4 + j, vals[j] if vals[j] else None, fmt=NUM,
               align="right", size=8)
        sumrow(r)
        fin_idx[it] = r
        r += 1
    ws.merge_cells(start_row=fin_start, start_column=1,
                   end_row=r - 1, end_column=1)
    _w(ws, fin_start, 1, "재무\n활동", bold=True)
    _w(ws, r, 1, "재무 활동 자금수지", fill=GOLD, bold=True)
    _fillrow(ws, r, 1, 3, GOLD)
    for j in range(n):
        f = (f"={L(j)}{fin_idx['차입금 발생']}-{L(j)}{fin_idx['차입금 상환']}"
             f"-{L(j)}{fin_idx['현금 배당 지급']}")
        _w(ws, r, 4 + j, f, fill=GOLD, bold=True, fmt=NUM, align="right",
           size=8)
    sumrow(r)
    ws.cell(row=r, column=tcol).fill = GOLD
    r_fin = r
    r += 1
    # 차입금 잔액
    _w(ws, r, 1, "차입금 잔액", fill=LBLUE, bold=True)
    _fillrow(ws, r, 1, 3, LBLUE)
    for j in range(n):
        prev = (str(loan_bal_now)
                if j == 0 else f"{L(j-1)}{r}")
        f = (f"={prev}+{L(j)}{fin_idx['차입금 발생']}"
             f"-{L(j)}{fin_idx['차입금 상환']}")
        _w(ws, r, 4 + j, f, fill=LBLUE, bold=True, fmt=NUM, align="right",
           size=8)
    _w(ws, r, tcol, f"={L(n-1)}{r}", fill=LBLUE, bold=True, fmt=NUM,
       align="right")
    r += 1

    # ---- 투자활동 ----
    inv_start = r
    inv_idx = {}
    for it, vals in data["inv"].items():
        _w(ws, r, 2, it, size=8, align="left")
        _w(ws, r, 3, "")
        for j in range(n):
            _w(ws, r, 4 + j, vals[j] if vals[j] else None, fmt=NUM,
               align="right", size=8)
        sumrow(r)
        inv_idx[it] = r
        r += 1
    ws.merge_cells(start_row=inv_start, start_column=1,
                   end_row=r - 1, end_column=1)
    _w(ws, inv_start, 1, "투자\n활동", bold=True)
    _w(ws, r, 1, "투자 활동 자금수지", fill=GOLD, bold=True)
    _fillrow(ws, r, 1, 3, GOLD)
    items = list(inv_idx.keys())
    for j in range(n):
        f = (f"={L(j)}{inv_idx[items[0]]}-{L(j)}{inv_idx[items[1]]}"
             f"-{L(j)}{inv_idx[items[2]]}")
        _w(ws, r, 4 + j, f, fill=GOLD, bold=True, fmt=NUM, align="right",
           size=8)
    sumrow(r)
    ws.cell(row=r, column=tcol).fill = GOLD
    r_inv = r
    r += 1

    # ---- 월 수지 / 기말 잔액 ----
    _w(ws, r, 1, "월 자금수지", fill=GOLD, bold=True)
    _fillrow(ws, r, 1, 3, GOLD)
    for j in range(n):
        _w(ws, r, 4 + j, f"={L(j)}{r_op}+{L(j)}{r_fin}+{L(j)}{r_inv}",
           fill=GOLD, bold=True, fmt=NUM, align="right", size=8)
    sumrow(r)
    ws.cell(row=r, column=tcol).fill = GOLD
    r_net = r
    r += 1
    _w(ws, r, 1, "기말 잔액", fill=GRAYF, bold=True)
    _fillrow(ws, r, 1, 3, GRAYF)
    for j in range(n):
        _w(ws, r, 4 + j, f"={L(j)}{r_base}+{L(j)}{r_net}",
           fill=GRAYF, bold=True, fmt=NUM, align="right", size=8)
    _w(ws, r, tcol, f"={L(n-1)}{r}", fill=GRAYF, bold=True, fmt=NUM,
       align="right")
    r_end = r
    # 기초 시재 연동 수식 (2번째 달부터 = 전월 기말)
    for j in range(1, n):
        _w(ws, r_base, 4 + j, f"={L(j-1)}{r_end}", fmt=NUM, align="right",
           bold=True, fill=GRAYF, size=8)
    _w(ws, r_base, tcol, f"={L(0)}{r_base}", fmt=NUM, align="right",
       bold=True, fill=GRAYF)


def build_cashplan_excel(months, current_ym, opening, data,
                         loan_bal_now, summary) -> bytes:
    wb = Workbook()
    sheet_summary(wb, summary)
    sheet_cashflow(wb, months, current_ym, opening, data, loan_bal_now)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
