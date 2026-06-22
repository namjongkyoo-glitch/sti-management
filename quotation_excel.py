"""고객 제출용 견적서 엑셀 생성 (갑지 / 속지 / Labor & Materials)"""
import os
from io import BytesIO
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sti_logo.png")


def _logo(ws, anchor: str, height: int = 42):
    """로고 삽입 (파일이 없거나 Pillow 미설치 시 조용히 생략)"""
    try:
        img = XLImage(LOGO_PATH)
        scale = height / img.height
        img.height = height
        img.width = int(img.width * scale)
        ws.add_image(img, anchor)
    except Exception:
        pass


# ---- 회사 정보 (필요 시 여기만 수정) ----
COMPANY_NAME = "STI AD USA INC"
COMPANY_ADDR = "2261 Gattis School Rd, Unit# 100, Round Rock, TX"
COMPANY_TEL = "Tel : 512-999-4442"
COMPANY_REP = "Jaden Nam, Branch Manager"

CYAN = PatternFill("solid", fgColor="DAEEF3")
GREEN = PatternFill("solid", fgColor="C6E0B4")
LGREEN = PatternFill("solid", fgColor="E2EFDA")
PEACH = PatternFill("solid", fgColor="FCE4D6")
TAN = PatternFill("solid", fgColor="FFF2CC")
BLUEGRAY = PatternFill("solid", fgColor="D6DCE4")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
USD = '"$" #,##0.00'


def _s(ws, r, c, v=None, *, fill=None, bold=False, size=None, color=None,
       align="center", fmt=None, border=True, underline=False, italic=False):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    if border:
        cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size or 11, color=color,
                     underline="single" if underline else None, italic=italic)
    if fmt:
        cell.number_format = fmt
    return cell


def _row(ws, r, c1, c2, *, fill=None):
    for c in range(c1, c2 + 1):
        cell = ws.cell(row=r, column=c)
        cell.border = BORDER
        if fill:
            cell.fill = fill


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# 갑지 (Quotation 표지)
# ============================================================
def sheet_cover(wb, est, customer=False):
    ws = wb.active
    ws.title = "갑지(Quotation)"
    _widths(ws, [38, 16, 10, 12, 12, 15, 14])

    ws.merge_cells("A1:G1")
    _s(ws, 1, 1, "Quotation", bold=True, size=24, border=False)
    ws.row_dimensions[1].height = 34

    _logo(ws, "A3", height=34)
    _s(ws, 3, 1, f"             {COMPANY_NAME}", bold=True, size=15,
       color="1F3864", align="left", border=False)
    ws.row_dimensions[3].height = 28
    _s(ws, 4, 1, COMPANY_ADDR, bold=True, size=9, align="left", border=False)
    _s(ws, 5, 1, COMPANY_TEL, bold=True, size=9, align="left", border=False)
    _s(ws, 6, 1, COMPANY_REP, bold=True, size=9, align="left", border=False)

    _s(ws, 8, 1, f"Project : {est['title']}", bold=True, align="left", border=False)
    _s(ws, 9, 1, f"Client : {est.get('client') or ''}    /    Date : {date.today()}",
       size=9, align="left", border=False)
    _s(ws, 10, 2, "Quotation valid for 30 days", bold=True,
       color="0000FF", align="left", border=False)

    hdr = ["Description", "규격", "Unit", "Material", "Labor", "Amount", "Remarks"]
    R = 12
    for i, h in enumerate(hdr, 1):
        _s(ws, R, i, h, fill=CYAN, bold=True, size=9)

    # ---- 품목 행: 직접비 항목 자동 입력 (없으면 빈 양식 3행) ----
    direct_items = est.get("_direct_items") or []
    order_amount = float(est.get("_order_amount") or 0)
    n_rows = max(3, len(direct_items))
    for k in range(n_rows):
        rr = R + 1 + k
        _row(ws, rr, 1, 7)
        if k < len(direct_items):
            _s(ws, rr, 1, direct_items[k]["item"], align="left", size=9)
            _s(ws, rr, 3, "LOT", size=9)
            _s(ws, rr, 6, direct_items[k]["amount"], fmt=USD, size=9)
        else:
            if k == 0:
                _s(ws, rr, 1, est["title"], align="left", size=9)
                _s(ws, rr, 3, "LOT", size=9)
            _s(ws, rr, 6, None, fmt=USD)

    r_direct = R + 1 + n_rows
    ws.merge_cells(start_row=r_direct, start_column=1, end_row=r_direct, end_column=5)
    _s(ws, r_direct, 1, "Direct Cost Total", fill=CYAN, bold=True, size=9)
    _row(ws, r_direct, 1, 7, fill=CYAN)
    _s(ws, r_direct, 6, f"=SUM(F{R+1}:F{r_direct-1})",
       fill=CYAN, bold=True, fmt=USD)

    direct_sum = sum(float(d["amount"] or 0) for d in direct_items)

    if customer:
        # 고객 제출용: 내부 원가(직접비/이익 구분) 숨기고 총액만 표시
        r_grand = r_direct + 1
        ws.merge_cells(start_row=r_grand, start_column=1,
                       end_row=r_grand, end_column=5)
        _s(ws, r_grand, 1, "Grand Total", fill=CYAN, bold=True, size=11)
        _row(ws, r_grand, 1, 7, fill=CYAN)
        if order_amount:
            _s(ws, r_grand, 6, order_amount, fill=CYAN, bold=True,
               color="0000FF", fmt=USD)
        else:
            _s(ws, r_grand, 6, f"=F{r_direct}", fill=CYAN, bold=True,
               color="0000FF", fmt=USD)
        ws.row_dimensions[r_grand].height = 20
        return

    # ---- (내부용) Profit: 수주금액 - 직접비, 비율 자동 계산 ----
    if order_amount and direct_sum:
        pct = (order_amount - direct_sum) / direct_sum * 100
        profit_label = f"Profit {pct:.0f}%"
        profit_value = order_amount - direct_sum
    else:
        profit_label = "Profit 15%"
        profit_value = None

    r_profit = r_direct + 1
    _row(ws, r_profit, 1, 7)
    _s(ws, r_profit, 2, profit_label, bold=True, size=9)
    _s(ws, r_profit, 3, "Lot", size=9)
    if profit_value is not None:
        _s(ws, r_profit, 6, profit_value, fmt=USD)
    else:
        _s(ws, r_profit, 6, f"=F{r_direct}*0.15", fmt=USD)

    r_ind = r_profit + 1
    ws.merge_cells(start_row=r_ind, start_column=1, end_row=r_ind, end_column=5)
    _s(ws, r_ind, 1, "Indirect Cost SubTotal", fill=PEACH, bold=True, size=9)
    _row(ws, r_ind, 1, 7, fill=PEACH)
    _s(ws, r_ind, 6, f"=F{r_profit}", fill=PEACH, bold=True, fmt=USD)

    r_grand = r_ind + 1
    ws.merge_cells(start_row=r_grand, start_column=1, end_row=r_grand, end_column=5)
    _s(ws, r_grand, 1, "Grand Total", fill=CYAN, bold=True, size=11)
    _row(ws, r_grand, 1, 7, fill=CYAN)
    _s(ws, r_grand, 6, f"=F{r_direct}+F{r_ind}", fill=CYAN, bold=True,
       color="0000FF", fmt=USD)
    ws.row_dimensions[r_grand].height = 20


# ============================================================
# 을지 (Management 요약)
# ============================================================
def sheet_mgmt(wb):
    ws = wb.create_sheet("을지(Management)")
    _widths(ws, [6, 32, 16, 12, 16, 32])

    ws.merge_cells("B1:E2")
    _s(ws, 1, 2, "MANAGEMENT", bold=True, size=22, underline=True, border=False)
    _s(ws, 1, 1, COMPANY_NAME, bold=True, size=9, color="1F3864", border=False)

    hdr = ["No.", "DESCRIPTION", "Unit Rate($/FT)", "Q'ty(FT)", "Amount", "Remark"]
    R = 4
    for i, h in enumerate(hdr, 1):
        _s(ws, R, i, h, fill=GREEN, bold=True, size=10)

    # 샘플 행 5개 (금액 = 단가 x 수량 자동 계산)
    for n in range(1, 6):
        rr = R + n
        _row(ws, rr, 1, 6)
        _s(ws, rr, 1, n if n <= 2 else None, fill=BLUEGRAY, size=9)
        _s(ws, rr, 2, None, fill=BLUEGRAY, size=9)
        _s(ws, rr, 3, None, fmt=USD, size=9)
        _s(ws, rr, 5, f"=IF(C{rr}*D{rr}=0,\"\",C{rr}*D{rr})", fmt=USD, size=9)

    r_tot = R + 7
    _row(ws, r_tot, 1, 6, fill=PEACH)
    _s(ws, r_tot, 2, "Total", fill=PEACH, bold=True)
    _s(ws, r_tot, 5, f"=SUM(E{R+1}:E{R+5})", fill=PEACH, bold=True, fmt=USD)


# ============================================================
# 속지 (Management Detail)
# ============================================================
def sheet_mgmt_detail(wb):
    ws = wb.create_sheet("속지(Detail)")
    _widths(ws, [6, 32, 8, 7, 7, 14, 15, 28])

    ws.merge_cells("B1:G2")
    _s(ws, 1, 2, "MANAGEMENT", bold=True, size=22, underline=True, border=False)
    _logo(ws, "A1", height=42)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    hdr = ["No.", "DESCRIPTION", "Unit", "Q'ty", "Q'ty", "Price", "Amount", "Remark"]
    R = 4
    for i, h in enumerate(hdr, 1):
        _s(ws, R, i, h, fill=GREEN, bold=True, size=10)

    r = R + 1
    subtotal_rows = []

    def section(no, name, items):
        nonlocal r
        _row(ws, r, 1, 8, fill=BLUEGRAY)
        _s(ws, r, 1, no, fill=BLUEGRAY, bold=True, size=10)
        _s(ws, r, 2, name, fill=BLUEGRAY, bold=True, size=10)
        r += 1
        start = r
        for it in items:
            _row(ws, r, 1, 8)
            _s(ws, r, 2, it[0], size=9)
            _s(ws, r, 3, it[1], size=9)
            _s(ws, r, 6, None, fmt=USD, size=9)
            _s(ws, r, 7, f"=IF(PRODUCT(D{r}:F{r})=0,\"\",PRODUCT(D{r}:F{r}))",
               fmt=USD, size=9)
            r += 1
        _row(ws, r, 1, 8)
        _s(ws, r, 2, "SUB TOTAL", bold=True, size=9)
        _s(ws, r, 7, f"=SUM(G{start}:G{r-1})", bold=True, fmt=USD, size=9)
        subtotal_rows.append(r)
        r += 1

    section(1, "Labor", [
        ("ON FIELD MANAGER", "Month"), ("Project Control Manager", "Month"),
        ("Quality management", "Month"), ("Safety management", "Month"),
        ("BIM, Drawing", "Month"), ("Overhead costs", "EA"),
    ])
    section(2, "Safety", [
        ("Provide safety gear/PPE/supplies", "Month"),
        ("OSHA Training", "Month"),
    ])
    section(3, "Subcontractor installation", [
        ("", "EA"), ("", "EA"), ("", "EA"),
    ])

    _row(ws, r, 1, 8, fill=PEACH)
    _s(ws, r, 2, "Total", fill=PEACH, bold=True)
    formula = "=" + "+".join(f"G{x}" for x in subtotal_rows)
    _s(ws, r, 7, formula, fill=PEACH, bold=True, fmt=USD)


# ============================================================
# Labor & Materials
# ============================================================
def sheet_labor_materials(wb):
    ws = wb.create_sheet("Labor & Materials")
    _widths(ws, [5, 36, 8, 7, 14, 15, 30])

    ws.merge_cells("B1:E2")
    c = _s(ws, 1, 2, "Labor & Materials", bold=True, size=20, underline=True)
    for rr in range(1, 3):
        for cc in range(2, 6):
            ws.cell(row=rr, column=cc).border = BORDER
    _logo(ws, "A1", height=42)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    hdr = ["No.", "DESCRIPTION", "Unit", "Q'ty", "Price", "Amount", "Remark"]
    R = 4
    for i, h in enumerate(hdr, 1):
        _s(ws, R, i, h, fill=TAN, bold=True, size=10)

    r = R + 1
    amount_rows = []

    def item_row(desc="", unit="", *, fill=None, bold=False, size=9, formula=True):
        nonlocal r
        _row(ws, r, 1, 7, fill=fill)
        _s(ws, r, 2, desc, fill=fill, bold=bold, size=size,
           align="left" if not bold else "center")
        _s(ws, r, 3, unit, fill=fill, size=9)
        _s(ws, r, 5, None, fill=fill, fmt=USD, size=9)
        if formula:
            _s(ws, r, 6, f"=IF(D{r}*E{r}=0,\"\",D{r}*E{r})",
               fill=fill, fmt=USD, size=9)
            amount_rows.append(r)
        r += 1

    # 1. 자재 섹션 (소항목 + Sub total)
    _row(ws, r, 1, 7, fill=LGREEN)
    _s(ws, r, 1, 1, fill=LGREEN, bold=True)
    _s(ws, r, 2, "TANK / MATERIAL (품명 입력)", fill=LGREEN, bold=True,
       size=11, align="left")
    r += 1
    sec_start = r
    for _ in range(3):
        item_row()
    sub_rows = amount_rows[-3:]
    _row(ws, r, 1, 7, fill=LGREEN)
    _s(ws, r, 2, "Sub total", fill=LGREEN, bold=True, size=9)
    _s(ws, r, 6, f"=SUM(F{sec_start}:F{r-1})", fill=LGREEN, bold=True, fmt=USD, size=9)
    sub_row = r
    for x in sub_rows:
        amount_rows.remove(x)
    amount_rows.append(sub_row)
    r += 1

    # 2~4 단일 항목
    for no, name in [(2, "Wood Packing"), (3, "Anchor Installation"), (4, "Shipping")]:
        _s(ws, r, 1, no, bold=True)
        item_row(name, "LOT", bold=True, size=10)

    r_tot = r + 1
    _row(ws, r_tot, 1, 7, fill=PEACH)
    _s(ws, r_tot, 2, "Total", fill=PEACH, bold=True)
    formula = "=" + "+".join(f"F{x}" for x in amount_rows)
    _s(ws, r_tot, 6, formula, fill=PEACH, bold=True, fmt=USD)


# ============================================================
# ============================================================
# 동적 속지 시트 (입력된 데이터로 생성)
# ============================================================
import re as _re


def sheet_detail_dynamic(wb, name: str, items: list[dict]):
    title = _re.sub(r"[\\/*?:\[\]]", "", name)[:31] or "Detail"
    ws = wb.create_sheet(title)
    _widths(ws, [6, 32, 8, 7, 7, 13, 15, 26])

    ws.merge_cells("B1:G2")
    _s(ws, 1, 2, name, bold=True, size=20, underline=True, border=False)
    _logo(ws, "A1", height=42)
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    hdr = ["No.", "DESCRIPTION", "Unit", "Q'ty", "Q'ty", "Price",
           "Amount", "Remark"]
    R = 4
    for i, h in enumerate(hdr, 1):
        _s(ws, R, i, h, fill=GREEN, bold=True, size=10)

    # 섹션별 그룹화 (입력 순서 유지)
    groups, order = {}, []
    for it in items:
        sec = (it.get("section") or "").strip()
        if sec not in groups:
            groups[sec] = []
            order.append(sec)
        groups[sec].append(it)

    r = R + 1
    sub_rows = []
    sec_no = 0
    for sec in order:
        has_section = bool(sec)
        if has_section:
            sec_no += 1
            m = _re.match(r"^(\d+)\.?\s*(.*)$", sec)
            no_txt, sec_name = (m.group(1), m.group(2)) if m else (str(sec_no), sec)
            _row(ws, r, 1, 8, fill=BLUEGRAY)
            _s(ws, r, 1, no_txt, fill=BLUEGRAY, bold=True, size=10)
            _s(ws, r, 2, sec_name, fill=BLUEGRAY, bold=True, size=10)
            r += 1
        start = r
        for it in groups[sec]:
            _row(ws, r, 1, 8)
            _s(ws, r, 2, it.get("description") or "", size=9, align="left")
            _s(ws, r, 3, it.get("unit") or "", size=9)
            q1, q2 = float(it.get("qty1") or 0), float(it.get("qty2") or 0)
            _s(ws, r, 4, q1 if q1 else None, size=9)
            _s(ws, r, 5, q2 if q2 else None, size=9)
            p = float(it.get("price") or 0)
            _s(ws, r, 6, p if p else None, fmt=USD, size=9)
            _s(ws, r, 7, f"=IF(PRODUCT(D{r}:F{r})=0,\"\",PRODUCT(D{r}:F{r}))",
               fmt=USD, size=9)
            _s(ws, r, 8, it.get("remark") or "", size=8, align="left")
            r += 1
        _row(ws, r, 1, 8)
        _s(ws, r, 2, "SUB TOTAL" if has_section else "Sub total",
           bold=True, size=9)
        _s(ws, r, 7, f"=SUM(G{start}:G{r-1})", bold=True, fmt=USD, size=9)
        sub_rows.append(r)
        r += 1

    _row(ws, r, 1, 8, fill=PEACH)
    _s(ws, r, 2, "Total", fill=PEACH, bold=True)
    if sub_rows:
        _s(ws, r, 7, "=" + "+".join(f"G{x}" for x in sub_rows),
           fill=PEACH, bold=True, fmt=USD)


def build_quotation_excel(est: dict, customer: bool = False) -> bytes:
    wb = Workbook()
    sheet_cover(wb, est, customer=customer)
    # 고객 제출용은 갑지(총액)만, 내부용은 속지 상세 + 별첨까지 포함
    if not customer:
        sheets = est.get("_sheets") or []
        if sheets:
            for s in sheets:
                sheet_detail_dynamic(wb, s["name"], s["items"])
        else:
            sheet_mgmt_detail(wb)
            sheet_labor_materials(wb)
        # 별첨 (제작비용/직접경비/현지운영비)
        if any(est.get(k) for k in ("sheet1_data", "sheet2_data", "sheet3_data")):
            from proposal_excel import _build_attachments
            _build_attachments(wb, est)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
