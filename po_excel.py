"""발주서(Purchase Order) 엑셀 — STI AD USA INC 양식"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quotation_excel import _logo

GRAY = PatternFill("solid", fgColor="D9DDE3")
THIN = Side(style="thin")
MED = Side(style="medium")
USD = '#,##0.00'

CO_NAME = "STI AD USA, INC."
CO_ADDR = "2261 Gattis School Rd #100, Round Rock, TX 78664"
CO_TEL = "TEL:1-512-999-4442"


def _c(ws, r, c, v=None, *, bold=False, size=10, align="left", fill=None,
       fmt=None, border=None, color=None):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    cell.font = Font(bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = border
    return cell


def build_po_excel(po: dict, lines: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    widths = [8, 30, 8, 7, 14, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- 헤더: 로고 + 회사정보 ----
    _logo(ws, "A1", height=34)
    _c(ws, 1, 5, CO_NAME, bold=True, size=10, align="right")
    ws.merge_cells("E1:G1")
    _c(ws, 2, 5, CO_ADDR, size=8, align="right")
    ws.merge_cells("E2:G2")
    _c(ws, 3, 5, CO_TEL, size=8, align="right")
    ws.merge_cells("E3:G3")

    # ---- 제목 ----
    ws.merge_cells("A5:G5")
    _c(ws, 5, 1, "Purchase Order", bold=True, size=20, align="center")
    ws.row_dimensions[5].height = 30

    # ---- 공급자 / PO 정보 ----
    r = 7
    _c(ws, r, 1, "Supplier's Name & Address", bold=True, size=9)
    _c(ws, r, 4, "PO Date", bold=True, size=9)
    _c(ws, r, 5, str(po.get("po_date") or "")[:10], size=9)
    _c(ws, r + 1, 1, po.get("supplier_name") or "", bold=True, size=10)
    _c(ws, r + 1, 4, "PO No", bold=True, size=9)
    _c(ws, r + 1, 5, po.get("po_no") or "", size=9)
    _c(ws, r + 2, 1, po.get("supplier_address") or "", size=8)
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=3)
    _c(ws, r + 2, 4, "PJT No", bold=True, size=9)
    _c(ws, r + 2, 5, po.get("pjt_no") or "", size=9)

    r = 11
    _c(ws, r, 1, "Attn.", bold=True, size=9)
    _c(ws, r, 2, po.get("attn") or "", bold=True, size=9)
    _c(ws, r, 4, "Order by", bold=True, size=9)
    _c(ws, r, 5, po.get("order_by") or "", size=9)
    _c(ws, r + 1, 1, "TEL", bold=True, size=9)
    _c(ws, r + 1, 2, po.get("supplier_tel") or "", size=9)
    _c(ws, r + 1, 4, "TEL", bold=True, size=9)
    _c(ws, r + 1, 5, po.get("order_by_tel") or "", size=9)
    _c(ws, r + 2, 1, "Email", bold=True, size=9)
    _c(ws, r + 2, 2, po.get("supplier_email") or "", bold=True, size=9,
       color="0000FF")
    _c(ws, r + 2, 4, "FAX", bold=True, size=9)

    # 구분선
    r = 14
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = Border(bottom=MED)

    _c(ws, 15, 1,
       "We have pleasure of ordering you the undermentioned goods, "
       "subject to the terms and conditions as follows.", bold=True, size=9)
    ws.merge_cells("A15:G15")

    # ---- 1~9 조건 ----
    r = 16
    conds = [
        ("1.Destination :", po.get("destination") or "",
         "2.Payment Terms", po.get("payment_terms") or ""),
        ("   " + (po.get("destination_addr") or ""), "", "", ""),
        ("3.Loading Port :", po.get("loading_port") or "",
         "4.Incoterms:", po.get("incoterms") or ""),
        ("5.Ship Mode :", po.get("ship_mode") or "",
         "6.Agent :", po.get("agent") or "-"),
        ("7.Packing :", po.get("packing") or "",
         "8.Shipping Type :", po.get("shipping_type") or ""),
        ("9.Remark :", po.get("remark") or "", "", ""),
    ]
    for left, lv, right, rv in conds:
        _c(ws, r, 1, f"{left} {lv}".strip(), bold=left.endswith(":"), size=9)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        if right:
            _c(ws, r, 4, f"{right} {rv}".strip(), bold=True, size=9)
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        r += 1

    _c(ws, r, 6, "Currency :  USD", bold=True, size=9, align="right")
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
    r += 1

    # ---- 품목 헤더 ----
    box = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    heads = ["Item No", "Description\nSpec", "Unit", "Qty",
             "Unit Price", "Amount", "Remark"]
    for i, h in enumerate(heads, 1):
        _c(ws, r, i, h, bold=True, size=9, align="center", fill=GRAY,
           border=box)
    r += 1
    first = r
    total = 0.0
    for idx, l in enumerate(lines, 1):
        qty = float(l.get("qty") or 0)
        price = float(l.get("unit_price") or 0)
        _c(ws, r, 1, idx, size=9, align="center", border=box)
        _c(ws, r, 2, l.get("description") or "", bold=True, size=8,
           border=box)
        _c(ws, r, 3, l.get("unit") or "LOT", size=9, align="center",
           border=box)
        _c(ws, r, 4, qty, size=9, align="center", border=box,
           fmt='#,##0')
        _c(ws, r, 5, price, size=9, align="right", border=box, fmt=USD)
        _c(ws, r, 6, f"=D{r}*E{r}", size=9, align="right", border=box,
           bold=True, fmt=USD)
        _c(ws, r, 7, l.get("remark") or "", size=8, align="center",
           border=box)
        total += qty * price
        r += 1

    # ---- 합계 ----
    _c(ws, r, 5, "Total Amount", bold=True, size=9, align="right")
    _c(ws, r, 6, f"=SUM(F{first}:F{r-1})", bold=True, size=10,
       align="right", fmt=USD)
    r += 2

    # ---- Remark / 약관 ----
    _c(ws, r, 1, "*Remark", bold=True, size=8)
    r += 1
    terms = [
        "1. Supplier herein shall indemnify the buyer for all claims of "
        "patent infringement and for all costs for defending against claims "
        "resulting from purchase of the above items.",
        "2. Supplier must give a notice prior to delivery to buyer if any of "
        "the items in this P/O is Under Any regulation or restriction.",
        "3. If you have no different opinion with this P/O, then you should "
        "put signature to supplier's empty column at bottom and reply to "
        "STI AD USA, INC.",
    ]
    for t in terms:
        _c(ws, r, 1, t, size=7)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
