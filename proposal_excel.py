"""품의서 엑셀 출력 (STI QSP-401-08 양식)"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from quotation_excel import _logo

GREEN = PatternFill("solid", fgColor="A9D08E")
PINK = PatternFill("solid", fgColor="F7A8C9")
YELLOW = PatternFill("solid", fgColor="FFF2A0")
GRAY = PatternFill("solid", fgColor="D9D9D9")
HEAD = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _c(ws, r, c, v=None, *, fill=None, bold=False, size=10, align="center",
       color=None, fmt=None, border=True, underline=False):
    cell = ws.cell(row=r, column=c)
    if v is not None:
        cell.value = v
    if border:
        cell.border = BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=size, color=color,
                     underline="single" if underline else None)
    if fmt:
        cell.number_format = fmt
    return cell


def _box(ws, r1, c1, r2, c2, fill=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER
            if fill:
                ws.cell(row=r, column=c).fill = fill


def build_proposal_excel(p: dict, author: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "품의서"
    # 컬럼: A 라벨(12) B 값(34) C~F 결재(9)
    for i, w in enumerate([12, 34, 9, 9, 9, 9], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    cur = "원" if p.get("currency") == "KRW" else "USD"
    NUM = '#,##0'

    # ---- 제목 ----
    _c(ws, 1, 1, "품 의 서", bold=True, size=20, align="left",
       border=False, underline=True)
    _logo(ws, "F1", height=30)
    ws.row_dimensions[1].height = 28

    # ---- 헤더 표 ----
    r = 3
    _c(ws, r, 1, "문서 번호", fill=HEAD, bold=True)
    _c(ws, r, 2, p.get("doc_no") or "", align="left")
    for i, h in enumerate(["담 당", "팀 장", "임 원", "CEO"], 3):
        _c(ws, r, i, h, fill=HEAD, bold=True, size=9)
    _c(ws, r + 1, 1, "보존 년한", fill=HEAD, bold=True)
    _c(ws, r + 1, 2, "( 3 ) 년", align="left")
    _c(ws, r + 2, 1, "작 성 일", fill=HEAD, bold=True)
    _c(ws, r + 2, 2, str(p.get("created_at") or "")[:10], align="left")
    _c(ws, r + 3, 1, "작 성 자", fill=HEAD, bold=True)
    _c(ws, r + 3, 2, f"미국법인 {author}", align="left")
    for i in range(3, 7):  # 결재 사인란
        ws.merge_cells(start_row=r + 1, start_column=i,
                       end_row=r + 3, end_column=i)
        _box(ws, r + 1, i, r + 3, i)
    _c(ws, r + 4, 1, "사본 배포", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r + 4, start_column=2, end_row=r + 4, end_column=6)
    _c(ws, r + 4, 2, p.get("copy_to") or "", align="left")
    _box(ws, r + 4, 2, r + 4, 6)
    _c(ws, r + 5, 1, "의      견", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r + 5, start_column=2, end_row=r + 5, end_column=6)
    _c(ws, r + 5, 2, p.get("opinion") or "", align="left")
    _box(ws, r + 5, 2, r + 5, 6)
    ws.row_dimensions[r + 5].height = 24
    _c(ws, r + 6, 1, "제      목", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r + 6, start_column=2, end_row=r + 6, end_column=6)
    _c(ws, r + 6, 2, f"{p.get('title') or ''} ({p.get('proposal_type')})",
       align="left", bold=True)
    _box(ws, r + 6, 2, r + 6, 6)

    # ---- 본문 ----
    r += 8
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=6)
    body = p.get("body") or (
        f"아래와 같이 {p.get('project_name') or ''}\n"
        "공사를 진행하고자 하오니 재가하여 주시기 바랍니다.")
    _c(ws, r, 1, body, border=False)
    ws.row_dimensions[r].height = 18
    ws.row_dimensions[r + 1].height = 18

    # ---- 프로젝트 정보 ----
    r += 3
    info = [
        ("1. Project  명 :", p.get("project_name")),
        ("2. Project  NO :", p.get("project_no")),
        ("3. 고  객  명 :", p.get("client")),
        ("4. 공  사  명 :", p.get("work_name")),
        ("5. 수      량 :", p.get("qty")),
        ("6. 수 주 금 액 :",
         f"{float(p.get('order_amount') or 0):,.0f} {cur}"
         + (f"  ({p.get('order_amount_note')})" if p.get("order_amount_note") else "")),
        ("7. 결 제 조 건 :", p.get("payment_terms")),
        ("8. P / O 여부 :", p.get("po_info")),
        ("9. 납 기 일 자 :", p.get("delivery_date")),
        ("10. 집 행 내 역 :", ""),
    ]
    for label, val in info:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        _c(ws, r, 1, label, align="left", border=False, size=10)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        _c(ws, r, 3, val or "", align="left", border=False, size=10)
        ws.row_dimensions[r].height = 16
        r += 1

    # ---- 집행내역 원가표 ----
    r += 1
    order = float(p.get("order_amount") or 0)
    profit = float(p.get("target_profit") or 0)
    mat, out, dexp = (float(p.get("material_cost") or 0),
                      float(p.get("outsourcing_cost") or 0),
                      float(p.get("direct_expense") or 0))
    lab, mfg, sga = (float(p.get("labor_cost") or 0),
                     float(p.get("mfg_overhead") or 0),
                     float(p.get("sga_cost") or 0))
    res = float(p.get("reserve") or 0)
    c_sub, d_sub = mat + out + dexp, lab + mfg + sga
    total = c_sub + d_sub

    # 별첨 시트 먼저 생성 -> 합계 셀 참조 확보 (품의서 본문이 링크)
    refs = _build_attachments(wb, p)

    def pct(x):
        return f"{x / order * 100:.0f}%" if order else "-"

    def dpct(x):
        return f"직접비의 {x / c_sub * 100:.0f}%" if c_sub else ""

    # 헤더 2단
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    _c(ws, r, 1, "구 분", fill=HEAD, bold=True)
    ws.cell(row=r + 1, column=1).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    _c(ws, r, 2, "항 목", fill=HEAD, bold=True)
    _c(ws, r + 1, 2, "대분류", fill=HEAD, bold=True, size=9)
    _c(ws, r + 1, 3, "중분류", fill=HEAD, bold=True, size=9)
    # 항목 칸이 좁아 B열을 대분류, C열(결재칸 첫째) 사용 -> C열 폭 보정
    ws.column_dimensions["C"].width = 13
    ws.merge_cells(start_row=r, start_column=4, end_row=r + 1, end_column=4)
    _c(ws, r, 4, f"금액 ({cur})", fill=HEAD, bold=True)
    ws.column_dimensions["D"].width = 14
    ws.merge_cells(start_row=r, start_column=5, end_row=r + 1, end_column=5)
    _c(ws, r, 5, "비율(%)", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r, start_column=6, end_row=r + 1, end_column=6)
    _c(ws, r, 6, "비 고", fill=HEAD, bold=True)
    ws.column_dimensions["F"].width = 16
    r += 2

    def line(c1, c2, amt, ratio, note="", fill=None, bold=False,
             merge12=False):
        nonlocal r
        if merge12:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            _c(ws, r, 1, c1, fill=fill, bold=bold, align="left", size=9)
            _box(ws, r, 1, r, 3, fill)
        else:
            _c(ws, r, 1, c1, fill=fill, size=9)
            _c(ws, r, 2, c2, fill=fill, size=9)
            _c(ws, r, 3, "", fill=fill)
        # amt: 숫자면 값, 문자열(수식 '=..')이면 그대로
        if isinstance(amt, str):
            val = amt
        else:
            val = amt if amt else ("-" if amt == 0 else amt)
        _c(ws, r, 4, val, fill=fill, bold=bold, fmt=NUM, align="right", size=9)
        _c(ws, r, 5, ratio, fill=fill, bold=bold, size=9)
        _c(ws, r, 6, note, fill=fill, size=8, align="left")
        r += 1

    # 수주액
    _c(ws, r, 1, "수주액", bold=True, size=9)
    _c(ws, r, 2, "수주 금액", size=9); _c(ws, r, 3, "")
    _c(ws, r, 4, order, fmt=NUM, align="right", size=9)
    _c(ws, r, 5, "100%", size=9); _c(ws, r, 6, "", size=8)
    order_row = r
    r += 1
    line("(A) 수주 금액 계", "", order, "100%", fill=GREEN, bold=True, merge12=True)
    line("(B) 이익 목표 금액", "", profit, pct(profit), fill=PINK, bold=True, merge12=True)

    cost_top = r
    # 재료비/외주비: 별첨1 참조, 직접경비: 별첨2 참조 (수식)
    mat_v = f"={refs['mat']}" if refs.get("mat") else mat
    out_v = f"={refs['out']}" if refs.get("out") else out
    dexp_v = f"={refs['dexp']}" if refs.get("dexp") else dexp

    line("총원가", "재료비", mat_v, pct(mat))
    mat_row = r - 1
    line("직접비", "외주비", out_v, pct(out))
    out_row = r - 1
    line("", "직접경비", dexp_v, pct(dexp))
    dexp_row = r - 1
    # (C) 소계 = 재료비+외주비+직접경비 (SUM)
    c_formula = f"=SUM(D{mat_row}:D{dexp_row})"
    line("(C) 소계", "", c_formula, pct(c_sub), fill=YELLOW, bold=True, merge12=True)
    c_row = r - 1
    line("간접비", "노무비", lab, pct(lab), dpct(lab))
    lab_row = r - 1
    line("(공통비)", "제조간접경비", mfg, pct(mfg), dpct(mfg))
    line("", "판관비", sga, pct(sga), dpct(sga))
    sga_row = r - 1
    # (D) 소계 = 노무비+제조간접+판관비 (SUM)
    d_formula = f"=SUM(D{lab_row}:D{sga_row})"
    line("(D) 소계", "", d_formula, pct(d_sub), fill=YELLOW, bold=True, merge12=True)
    d_row = r - 1
    # (E) 총원가 = C소계 + D소계
    line("(E) 총원가 계(C+D)", "", f"=D{c_row}+D{d_row}", pct(total),
         "수주액 대비", fill=GREEN, bold=True, merge12=True)
    e_row = r - 1
    line("(F) 예비비", "", res, f"{res/order*100:.1f}%" if order else "-",
         fill=GREEN, bold=True, merge12=True)
    f_row = r - 1
    # (G) 총예정원가 = E + F
    line("(G) 총예정원가( E + F = A - B )", "", f"=D{e_row}+D{f_row}",
         pct(total + res), "수주액 대비", fill=GRAY, bold=True, merge12=True)

    # ── 셀 병합: 직접비(재료~직접경비), 간접비(노무~판관비) ──
    # 직접비: A열 mat_row~dexp_row 병합
    ws.merge_cells(start_row=mat_row, start_column=1,
                   end_row=dexp_row, end_column=1)
    _c(ws, mat_row, 1, "직접비", bold=True, size=9, align="center")
    # 간접비(공통비): A열 lab_row~sga_row 병합
    ws.merge_cells(start_row=lab_row, start_column=1,
                   end_row=sga_row, end_column=1)
    _c(ws, lab_row, 1, "간접비\n(공통비)", bold=True, size=9, align="center")
    # 총원가 라벨은 직접비+간접비 전체를 감싸지 않고 직접비 위에만 표시되도록
    # (별도 '총원가' 셀은 제거: 직접비/간접비로 구분)
    # 예비비: 값이 없으면 0으로 표시 (line에서 res=0이면 이미 0)
    if not res:
        _c(ws, f_row, 4, 0, fmt=NUM, align="right", size=9,
           fill=GREEN, bold=True)

    # 푸터
    r += 2
    _c(ws, r, 1, "QSP-401-08(Rev.1)", border=False, size=8, align="left")
    _c(ws, r, 3, "STI Co., Ltd.", border=False, size=8)
    _c(ws, r, 6, "A4(210mm x 297mm)", border=False, size=8, align="right")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_attachments(wb, p):
    """별첨1/2/3 시트 생성. 합계 셀 참조(dict) 반환:
    {mat, out, dexp, local} 각각 '시트명'!셀 형태 (없으면 None)"""
    import proposal_sheets as ps
    s1 = p.get("sheet1_data") or {}
    s2 = p.get("sheet2_data") or []
    s3 = p.get("sheet3_data") or []
    refs = {"mat": None, "out": None, "dexp": None, "local": None}

    # ── 별첨1: 제작비용 내역 ──
    ws = wb.create_sheet("별첨1-제작비용내역")
    SN1 = "'별첨1-제작비용내역'"
    for i, w in enumerate([6, 14, 30, 8, 14, 16, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # 제목 A1:G1 병합
    ws.merge_cells("A1:G1")
    _c(ws, 1, 1, "별첨#1: 제작 비용 내역", bold=True, size=13, align="left",
       border=False)
    hdr = ["구분", "대분류", "중분류(품목)", "수량", "단가", "합계금액", "비고"]
    for c, h in enumerate(hdr, 1):
        _c(ws, 3, c, h, fill=HEAD, bold=True)
    r = 4
    mat_rows = s1.get("material", [])
    out_rows = s1.get("outsource", [])
    MINROWS = 5  # 데이터 없어도 최소 5줄 공백

    def make_block(label, rows, ref_key):
        nonlocal r
        # 빈 행 패딩으로 최소 MINROWS 보장
        display = list(rows) + [{} for _ in range(max(0, MINROWS - len(rows)))]
        start = r
        for i, row in enumerate(display):
            _c(ws, r, 1, label if i == 0 else "", align="center")
            _c(ws, r, 2, row.get("대분류", ""), align="left")
            _c(ws, r, 3, row.get("중분류", ""), align="left")
            qty = row.get("수량", "")
            _c(ws, r, 4, float(qty) if qty not in ("", None) else "")
            up = row.get("단가", "")
            _c(ws, r, 5, float(up) if up not in ("", None) else "", fmt="#,##0")
            # 합계금액 = 수량*단가 (빈 행도 수식, 빈 값이면 0)
            _c(ws, r, 6, f"=IF(AND(D{r}<>\"\",E{r}<>\"\"),D{r}*E{r},0)", fmt="#,##0")
            _c(ws, r, 7, row.get("비고", ""), align="left")
            r += 1
        # 합계 행
        _c(ws, r, 1, f"{label} 합계", fill=YELLOW, bold=True, align="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        _c(ws, r, 6, f"=SUM(F{start}:F{r-1})", fill=YELLOW, bold=True,
           fmt="#,##0")
        _c(ws, r, 7, "")
        refs[ref_key] = f"{SN1}!F{r}"
        r += 1

    make_block("원재료", mat_rows, "mat")
    make_block("외주비", out_rows, "out")

    # TOTAL = 원재료합계 + 외주비합계
    _c(ws, r, 1, "TOTAL 합계", fill=GREEN, bold=True, align="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    parts = [ref.split("!")[1] for ref in (refs["mat"], refs["out"]) if ref]
    total_formula = "=" + "+".join(parts) if parts else 0
    _c(ws, r, 6, total_formula, fill=GREEN, bold=True, fmt="#,##0")
    _c(ws, r, 7, "")

    # ── 별첨2/별첨3: 항목별 금액 + SUM 합계 ──
    def expense_sheet(title, rows, ref_key):
        w = wb.create_sheet(title)
        sn = f"'{title}'"
        for i, wd in enumerate([22, 18, 40], 1):
            w.column_dimensions[get_column_letter(i)].width = wd
        w.merge_cells("A1:C1")
        _c(w, 1, 1, title.replace("별첨", "별첨#"), bold=True, size=13,
           align="left", border=False)
        for c, h in enumerate(["구 분", "금액", "비 고"], 1):
            _c(w, 3, c, h, fill=HEAD, bold=True)
        rr = 4
        start = rr
        for row in rows:
            _c(w, rr, 1, row.get("구분", ""), align="left")
            _c(w, rr, 2, float(row.get("금액") or 0), fmt="#,##0")
            _c(w, rr, 3, row.get("비고", ""), align="left")
            rr += 1
        _c(w, rr, 1, "합계", fill=GREEN, bold=True)
        _c(w, rr, 2, f"=SUM(B{start}:B{rr-1})", fill=GREEN, bold=True,
           fmt="#,##0")
        _c(w, rr, 3, "")
        refs[ref_key] = f"{sn}!B{rr}"

    if s2:
        expense_sheet("별첨2-직접경비내역", s2, "dexp")
    if s3:
        expense_sheet("별첨3-현지운영비", s3, "local")

    return refs
