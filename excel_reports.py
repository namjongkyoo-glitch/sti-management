"""견적 원가표 엑셀 출력 (본품의/변경품의 버전 비교 양식)"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import helpers

GREEN = PatternFill("solid", fgColor="A9D08E")
YELLOW = PatternFill("solid", fgColor="FFF2A0")
PINK = PatternFill("solid", fgColor="F7A8C9")
GRAY = PatternFill("solid", fgColor="D9D9D9")
HEAD = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUM_FMT = '#,##0;[Red](#,##0)'
PCT_FMT = '0%'


def _set(ws, row, col, value, *, fill=None, bold=False, fmt=None,
         align="center", font_color=None, size=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if fill:
        c.fill = fill
    if bold or font_color or size:
        c.font = Font(bold=bold, color=font_color, size=size or 11)
    if fmt:
        c.number_format = fmt
    return c


def build_estimate_excel(estimate: dict, versions: list[dict],
                         lines_map: dict[int, list[dict]]) -> bytes:
    """
    estimate  : estimates 행
    versions  : estimate_versions 행 목록 (version_no 순)
    lines_map : {version_id: [{'mid': 중분류명, 'item': 항목명, 'amount': 금액}, ...]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "견적원가"

    n_ver = len(versions)
    has_diff = n_ver >= 2
    # 컬럼: A구분 B대분류 C중분류/항목 | (금액,비율)*버전 | [증감] | 비고
    ver_start = 4
    diff_col = ver_start + n_ver * 2 if has_diff else None
    note_col = (diff_col + 1) if has_diff else ver_start + n_ver * 2

    # ---------- 제목/단위 ----------
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=note_col)
    t = ws.cell(row=1, column=1,
                value=f"견적 원가표 - {estimate['estimate_no']} {estimate['title']}")
    t.font = Font(bold=True, size=14)
    t.alignment = Alignment(horizontal="center")
    u = ws.cell(row=2, column=note_col, value="단위: USD")
    u.alignment = Alignment(horizontal="right")

    # ---------- 헤더 (3~4행) ----------
    r1, r2 = 3, 4
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    _set(ws, r1, 1, "구 분", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r1, start_column=2, end_row=r1, end_column=3)
    _set(ws, r1, 2, "항 목", fill=HEAD, bold=True)
    _set(ws, r2, 2, "대분류", fill=HEAD, bold=True)
    _set(ws, r2, 3, "중분류 / 항목", fill=HEAD, bold=True)

    for i, v in enumerate(versions):
        c0 = ver_start + i * 2
        ws.merge_cells(start_row=r1, start_column=c0, end_row=r1, end_column=c0 + 1)
        label = f"{v['version_label']}\n{v['version_date']}"
        _set(ws, r1, c0, label, fill=HEAD, bold=True)
        _set(ws, r2, c0, "금액", fill=HEAD, bold=True)
        _set(ws, r2, c0 + 1, "비율(%)", fill=HEAD, bold=True)
    if has_diff:
        ws.merge_cells(start_row=r1, start_column=diff_col, end_row=r2, end_column=diff_col)
        _set(ws, r1, diff_col, "증감", fill=HEAD, bold=True)
    ws.merge_cells(start_row=r1, start_column=note_col, end_row=r2, end_column=note_col)
    _set(ws, r1, note_col, "비 고", fill=HEAD, bold=True)

    # ---------- 데이터 집계 ----------
    order_amts = [float(v["order_amount"] or 0) for v in versions]

    def sum_mid(vidx: int, mid: str) -> float:
        vid = versions[vidx]["id"]
        return sum(float(l["amount"] or 0) for l in lines_map.get(vid, [])
                   if l["mid"] == mid)

    # 직접비 항목 목록 (모든 버전 합집합, 순서 유지)
    items_by_mid: dict[str, list[str]] = {m: [] for m in helpers.COST_MIDS_DIRECT}
    for v in versions:
        for l in lines_map.get(v["id"], []):
            if l["mid"] in items_by_mid and l["item"] not in items_by_mid[l["mid"]]:
                items_by_mid[l["mid"]].append(l["item"])

    def item_amt(vidx: int, mid: str, item: str) -> float:
        vid = versions[vidx]["id"]
        return sum(float(l["amount"] or 0) for l in lines_map.get(vid, [])
                   if l["mid"] == mid and l["item"] == item)

    def write_amounts(row, getter, *, fill=None, bold=False, pct=True,
                      pct_base=None):
        """getter(vidx) -> 금액. pct_base 없으면 해당 버전 수주금액 기준."""
        vals = []
        for i in range(n_ver):
            amt = getter(i)
            vals.append(amt)
            base = (pct_base[i] if pct_base else order_amts[i]) or 0
            _set(ws, row, ver_start + i * 2, amt if amt else 0,
                 fill=fill, bold=bold, fmt=NUM_FMT, align="right")
            if pct:
                _set(ws, row, ver_start + i * 2 + 1,
                     (amt / base) if base else 0,
                     fill=fill, bold=bold, fmt=PCT_FMT)
            else:
                _set(ws, row, ver_start + i * 2 + 1, "", fill=fill)
        if has_diff:
            d = vals[-1] - vals[0]
            _set(ws, row, diff_col, d, fill=fill, bold=True, fmt=NUM_FMT,
                 align="right", font_color="FF0000" if d < 0 else None)
        _set(ws, row, note_col, "", fill=fill)
        return vals

    row = 5

    # ---------- 수주금액 ----------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set(ws, row, 1, "수주금액", fill=GREEN, bold=True, align="left")
    write_amounts(row, lambda i: order_amts[i], fill=GREEN, bold=True)
    row += 1

    cost_top = row  # '총원가' 세로 병합 시작

    # ---------- 직접비 ----------
    direct_top = row
    for mid in helpers.COST_MIDS_DIRECT:
        items = items_by_mid[mid] or ["-"]
        mid_top = row
        for it in items:
            _set(ws, row, 3, it, align="left")
            write_amounts(row, lambda i, m=mid, t=it: item_amt(i, m, t))
            row += 1
        _set(ws, row, 3, "소계", fill=GRAY, bold=True)
        write_amounts(row, lambda i, m=mid: sum_mid(i, m), fill=GRAY, bold=True)
        row += 1
        ws.merge_cells(start_row=mid_top, start_column=2,
                       end_row=row - 1, end_column=2)
        _set(ws, mid_top, 2, mid, fill=GRAY)
        for rr in range(mid_top, row):
            ws.cell(row=rr, column=2).border = BORDER

    # (C) 소계
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    _set(ws, row, 2, "(C) 소계  [직접비]", fill=YELLOW, bold=True, align="left")
    c_vals = write_amounts(
        row, lambda i: sum(sum_mid(i, m) for m in helpers.COST_MIDS_DIRECT),
        fill=YELLOW, bold=True)
    row += 1

    # ---------- 간접비 ----------
    ind_top = row
    for mid in helpers.COST_MIDS_INDIRECT:
        _set(ws, row, 3, mid, align="left")
        write_amounts(row, lambda i, m=mid: sum_mid(i, m))
        row += 1
    ws.merge_cells(start_row=ind_top, start_column=2, end_row=row - 1, end_column=2)
    _set(ws, ind_top, 2, "간접비\n(공통비)", fill=GRAY)
    for rr in range(ind_top, row):
        ws.cell(row=rr, column=2).border = BORDER

    # (D) 소계
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    _set(ws, row, 2, "(D) 소계  [간접비]", fill=YELLOW, bold=True, align="left")
    d_vals = write_amounts(
        row, lambda i: sum(sum_mid(i, m) for m in helpers.COST_MIDS_INDIRECT),
        fill=YELLOW, bold=True)
    row += 1

    # '총원가' 세로 병합
    ws.merge_cells(start_row=cost_top, start_column=1, end_row=row - 1, end_column=1)
    _set(ws, cost_top, 1, "총원가", bold=True)
    for rr in range(cost_top, row):
        ws.cell(row=rr, column=1).border = BORDER

    # ---------- 총원가 계 / 영업이익 ----------
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set(ws, row, 1, "총원가 계 (C+D)", fill=GREEN, bold=True, align="left")
    write_amounts(row, lambda i: c_vals[i] + d_vals[i], fill=GREEN, bold=True)
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    _set(ws, row, 1, "영업이익", fill=PINK, bold=True, align="left")
    write_amounts(row, lambda i: order_amts[i] - c_vals[i] - d_vals[i],
                  fill=PINK, bold=True)
    row += 1

    # ---------- 열 너비 ----------
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    for i in range(n_ver):
        ws.column_dimensions[get_column_letter(ver_start + i * 2)].width = 13
        ws.column_dimensions[get_column_letter(ver_start + i * 2 + 1)].width = 8
    if has_diff:
        ws.column_dimensions[get_column_letter(diff_col)].width = 11
    ws.column_dimensions[get_column_letter(note_col)].width = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 월별 급여 / 지급 세부내역 엑셀 (미국법인 양식)
# ============================================================
PAY_HEAD = PatternFill("solid", fgColor="D9E1F2")
PAY_SUM = PatternFill("solid", fgColor="FFF2CC")
USD2 = '#,##0.00;[Red](#,##0.00)'


def build_payroll_excel(year: int, month: int, lines: list[dict]) -> bytes:
    """lines: [{'name','base_salary','per_diem','ot_amount','meal','vehicle',
               'telecom','transfer_fee','clawback','subtotal',
               'withholding_tax','net_pay','remarks'}, ...]"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}월 급여"

    cols = ["구 분", "급여", "일비", "OT", "중식", "차량지원", "통신비",
            "송금수수료", "초과금\n환수", "소 계", "원천세", "차감\n지급액", "비 고"]
    widths = [10, 12, 9, 11, 9, 10, 9, 11, 11, 12, 11, 12, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    t = ws.cell(row=1, column=1,
                value=f"미국법인 {month}월 급여 / 지급 세부내역")
    t.font = Font(bold=True, size=16)
    t.alignment = Alignment(horizontal="center")
    u = ws.cell(row=2, column=13, value="(단위:USD)")
    u.alignment = Alignment(horizontal="right")

    # 헤더 2단 (회사 지급 총액 병합)
    r1, r2 = 3, 4
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    _set(ws, r1, 1, "구 분", fill=PAY_HEAD, bold=True)
    ws.merge_cells(start_row=r1, start_column=2, end_row=r1, end_column=9)
    _set(ws, r1, 2, "회사 지급 총액", fill=PAY_HEAD, bold=True)
    for c in range(2, 10):
        _set(ws, r2, c, cols[c - 1], fill=PAY_HEAD, bold=True, fmt=None)
    for c in range(10, 14):
        ws.merge_cells(start_row=r1, start_column=c, end_row=r2, end_column=c)
        _set(ws, r1, c, cols[c - 1], fill=PAY_HEAD, bold=True)
        ws.cell(row=r2, column=c).border = BORDER

    keys = ["base_salary", "per_diem", "ot_amount", "meal", "vehicle",
            "telecom", "transfer_fee", "clawback", "subtotal",
            "withholding_tax", "net_pay"]
    row = 5
    for ln in lines:
        _set(ws, row, 1, ln["name"], bold=False)
        for i, k in enumerate(keys, start=2):
            v = float(ln.get(k) or 0)
            _set(ws, row, i, v if v else None, fmt=USD2, align="right",
                 bold=(k == "net_pay"),
                 font_color="FF0000" if (k == "clawback" and v < 0) else None)
        _set(ws, row, 13, ln.get("remarks") or "", align="left")
        ws.row_dimensions[row].height = 22
        row += 1

    # 합계
    _set(ws, row, 1, "합계", fill=PAY_SUM, bold=True)
    for i in range(2, 13):
        col = get_column_letter(i)
        _set(ws, row, i, f"=SUM({col}5:{col}{row-1})",
             fill=PAY_SUM, bold=True, fmt=USD2, align="right")
    _set(ws, row, 13, "", fill=PAY_SUM)

    # ---- 직원별 계좌 정보 ----
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    t2 = ws.cell(row=row, column=1, value="직원별 계좌 정보 (급여 이체용)")
    t2.font = Font(bold=True, size=12)
    row += 1
    bank_hdr = ["이름", "은행명", "계좌종류", "계좌번호", "Routing #", "Zelle"]
    for i, h in enumerate(bank_hdr, 1):
        _set(ws, row, i, h, fill=PAY_HEAD, bold=True)
    row += 1
    for ln in lines:
        _set(ws, row, 1, ln["name"])
        _set(ws, row, 2, ln.get("bank_name") or "", align="left")
        _set(ws, row, 3, ln.get("account_type") or "")
        _set(ws, row, 4, ln.get("account_no") or "", align="left")
        _set(ws, row, 5, ln.get("routing_no") or "", align="left")
        _set(ws, row, 6, ln.get("zelle") or "", align="left")
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# 프로젝트 스케줄 엑셀 (회사 양식: 계획/실적 간트 그리드)
# ============================================================
from datetime import date as _date, timedelta as _td

SCH_HEAD = PatternFill("solid", fgColor="C6E0B4")
PLAN_FILL = PatternFill("solid", fgColor="B4C7E7")   # 계획: 파랑
ACT_FILL = PatternFill("solid", fgColor="70AD47")    # 실적: 초록


def _to_date(v):
    if not v:
        return None
    if isinstance(v, _date):
        return v
    try:
        return _date.fromisoformat(str(v)[:10])
    except Exception:
        return None


def build_schedule_excel(project: dict, items: list[dict]) -> bytes:
    from quotation_excel import _logo, COMPANY_NAME
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # ---- 날짜 범위 (주 단위 버킷) ----
    dates = []
    for it in items:
        for k in ("plan_start", "plan_end", "actual_start", "actual_end"):
            d = _to_date(it.get(k))
            if d:
                dates.append(d)
    weeks = []
    if dates:
        start = min(dates) - _td(days=min(dates).weekday())  # 월요일
        end = max(dates)
        cur = start
        while cur <= end and len(weeks) < 80:
            weeks.append(cur)
            cur += _td(days=7)

    fixed = ["No", "공정명", "구분", "담당", "계획 시작", "계획 종료",
             "실적 시작", "실적 종료", "진행률", "비고"]
    nfix = len(fixed)
    widths = [4, 24, 10, 9, 11, 11, 11, 11, 8, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(len(weeks)):
        ws.column_dimensions[get_column_letter(nfix + 1 + j)].width = 3.2

    total_cols = nfix + max(len(weeks), 1)

    # ---- 제목 ----
    _logo(ws, "A1", height=36)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=min(total_cols, 12))
    t = ws.cell(row=1, column=2, value="PROJECT SCHEDULE")
    t.font = Font(bold=True, size=18, underline="single")
    t.alignment = Alignment(horizontal="center")
    info = ws.cell(row=2, column=2,
                   value=f"Project: {project.get('code','')} {project.get('name','')}"
                         f"   /   Client: {project.get('client') or '-'}"
                         f"   /   기간: {project.get('start_date') or '-'}"
                         f" ~ {project.get('end_date') or '-'}"
                         f"   /   출력일: {_date.today()}")
    info.font = Font(size=9, bold=True)
    # 범례
    _set(ws, 2, nfix + 1, "", fill=PLAN_FILL)
    ws.cell(row=2, column=nfix + 2, value="계획").font = Font(size=8)
    if len(weeks) > 4:
        _set(ws, 2, nfix + 4, "", fill=ACT_FILL)
        ws.cell(row=2, column=nfix + 5, value="실적").font = Font(size=8)

    # ---- 헤더 (4행: 월 / 5행: 주 시작일) ----
    r_month, r_head = 4, 5
    for i, h in enumerate(fixed, 1):
        ws.merge_cells(start_row=r_month, start_column=i,
                       end_row=r_head, end_column=i)
        _set(ws, r_month, i, h, fill=SCH_HEAD, bold=True, size=9)
        ws.cell(row=r_head, column=i).border = BORDER
    # 월 병합 헤더
    j = 0
    while j < len(weeks):
        k = j
        label = f"{weeks[j].year}.{weeks[j].month:02d}"
        while k + 1 < len(weeks) and \
                f"{weeks[k+1].year}.{weeks[k+1].month:02d}" == label:
            k += 1
        ws.merge_cells(start_row=r_month, start_column=nfix + 1 + j,
                       end_row=r_month, end_column=nfix + 1 + k)
        _set(ws, r_month, nfix + 1 + j, label, fill=SCH_HEAD, bold=True, size=8)
        for x in range(j, k + 1):
            ws.cell(row=r_month, column=nfix + 1 + x).border = BORDER
        j = k + 1
    for j, wk in enumerate(weeks):
        _set(ws, r_head, nfix + 1 + j, f"{wk.month}/{wk.day}",
             fill=SCH_HEAD, size=7)

    # ---- 데이터 행 ----
    row = r_head + 1
    for n, it in enumerate(items, 1):
        ps, pe = _to_date(it.get("plan_start")), _to_date(it.get("plan_end"))
        as_, ae = _to_date(it.get("actual_start")), _to_date(it.get("actual_end"))
        vals = [n, it.get("task_name") or "", it.get("category") or "",
                it.get("owner") or "",
                str(ps or ""), str(pe or ""), str(as_ or ""), str(ae or ""),
                f"{float(it.get('progress') or 0):.0f}%", it.get("notes") or ""]
        for i, v in enumerate(vals, 1):
            _set(ws, row, i, v, size=9,
                 align="left" if i in (2, 10) else "center")
        for j, wk in enumerate(weeks):
            we = wk + _td(days=6)
            cell = ws.cell(row=row, column=nfix + 1 + j)
            cell.border = BORDER
            if ps and pe and ps <= we and pe >= wk:
                cell.fill = PLAN_FILL
            if as_ and (ae or _date.today()) >= wk and as_ <= we:
                a_end = ae or _date.today()
                if as_ <= we and a_end >= wk:
                    cell.fill = ACT_FILL
        ws.row_dimensions[row].height = 18
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
